from lib2to3.pgen2.pgen import DFAState
from django.shortcuts import render
from django.contrib import messages
from urllib.request import Request
from django.http import HttpResponseNotAllowed
from django.shortcuts import redirect, render, get_object_or_404, redirect
from django.http import HttpResponse
from .forms import Form_FContas, Form_Gerais
from django.contrib.auth.decorators import login_required
from .models import Contas, Gerais
from openpyxl import load_workbook
import pandas as pd 
from django.db.models import Sum
from django.db.models import Q
from datetime import date, datetime
from fillpdf import fillpdfs
from fillpdf import extractfillpdf
import openpyxl 

saldo_mês_atual =  0
saldo_final_extrato_betel = 0
sal_final_betel = 0
total_total = 0
saldo_conta_extrato_sem_betel = 0
sal_final_con = 0
don_ent_value = 0
don_sai_value = 0
con_ent_value = 0
con_sai_value = 0
out_ent_value = 0
out_sai_value = 0
sal_final_don = 0
total_C_TE = 0
total_O = 0
total_G = 0
total_betel = 0

def home(request):
    return render(request, 'cp/home.html') 
   

@login_required
def editar_fc(request, id):
    contas = get_object_or_404(Contas, pk=id)
    form = Form_FContas(instance=contas)
 
    if request.method == 'POST':
        form = Form_FContas(request.POST, instance=contas)
         
        if form.is_valid():
            contas.save()
            messages.info(request, 'Editado com sucesso')
            return redirect('/contas')
   
        else:
            return render(request, 'cp/editar_fc.html', {'form':form ,'contas': contas})  


    return render(request, 'cp/editar_fc.html', {'form':form ,'contas': contas})  


@login_required
def deletar_fc(request, id):
    del_contas = get_object_or_404(Contas, pk=id)
   
    if request.method == 'POST':
        del_contas.delete()

        messages.info(request, 'Apagado com sucesso')
        return redirect('/contas')

    return render(request, 'cp/deletar_fc.html')         

@login_required
def add_gerais_contas(request):
    if request.method == 'POST':
        gerais = Form_Gerais(request.POST)
        
        if gerais.is_valid() :
            gerais.save()
            gerais = Form_Gerais()

            messages.info(request, 'Inserido com sucesso')
            return redirect('/view_gerais_contas')
 
    else:
        gerais = Form_Gerais()
    
    return render(request, 'cp/add_gerais_contas.html', {'gerais': gerais})   

@login_required
def editar_gerais_contas(request, id):
    editar_gerais = get_object_or_404(Gerais, pk=id)
    form = Form_Gerais(instance=editar_gerais)
 
    if request.method == 'POST':
        form = Form_Gerais(request.POST, instance=editar_gerais)
         
        if form.is_valid():
            editar_gerais.save()
            messages.info(request, 'Editado com sucesso')
            return redirect('/view_gerais_contas')
   
        else:
            return render(request, 'cp/editar_gerais_contas.html', {'form':form ,'editar_gerais': editar_gerais})  


    return render(request, 'cp/editar_gerais_contas.html', {'form':form ,'editar_gerais': editar_gerais})  

@login_required
def view_gerais_contas(request):
    gerais = Gerais.objects.all()
    return render(request, 'cp/view_gerais_contas.html', {'gerais': gerais})

@login_required
def deletar_gerais_contas(request):
    del_gerais = Gerais.objects.all()
   
    if request.method == 'POST':
        del_gerais.delete()

        messages.info(request, 'Apagado com sucesso')
        return redirect('/view_gerais_contas')

    return render(request, 'cp/deletar_gerais_contas.html')           


@login_required
def deletar_folha_de_contas (request):
    del_gerais = Contas.objects.all()
   
    if request.method == 'POST':
        del_gerais.delete()

        messages.info(request, 'Apagado com sucesso')
        return redirect('/contas')

    return render(request, 'cp/deletar_fc.html')  


@login_required
def relatório_mensal(request):
    global saldo_mês_atual
    global total_betel
    
    relatório = Gerais.objects.all()
    don_ent = Contas.objects.aggregate(total=Sum('donativos_Entrada'))
    don_sai = Contas.objects.aggregate(total=Sum('donativos_Saída'))
    con_ent = Contas.objects.aggregate(total=Sum('conta_Entrada'))
    con_sai = Contas.objects.aggregate(total=Sum('conta_Saída'))
    out_ent = Contas.objects.aggregate(total=Sum('outra_Entrada'))
    out_sai = Contas.objects.aggregate(total=Sum('outra_Saída'))
 
    don_ent_value1 = don_ent['total']
    if don_ent_value1:
        don_ent_value = float(don_ent_value1)
    else:
        don_ent_value = 0    
    don_sai_value1 = don_sai['total']
    if don_sai_value1:
        don_sai_value = float(don_sai_value1)
    else:
        don_sai_value = 0    
    con_ent_value1 = con_ent['total']
    if con_ent_value1:
        con_ent_value = float(con_ent_value1)
    else:
        con_ent_value = 0    
    con_sai_value1 = con_sai['total']
    if con_sai_value1:
        con_sai_value = float(con_sai_value1)
    else:
        con_sai_value = 0    
    out_ent_value1 = out_ent['total']
    if out_ent_value1:
        out_ent_value = float(out_ent_value1)
    else:
        out_ent_value = 0    
    out_sai_value1 = out_sai['total']
    if out_sai_value1:
        out_sai_value = float(out_sai_value1)
    else:
        out_sai_value = 0   

    saldo_final_extrato_mensal2 = Gerais.objects.values('saldo_Final_do_Extrato_Mensal').last()
    if saldo_final_extrato_mensal2 is None:
        saldo_final_extrato_mensal = 0   
    else:
        saldo_final_extrato_mensal1 = saldo_final_extrato_mensal2['saldo_Final_do_Extrato_Mensal']
        if saldo_final_extrato_mensal1:
            saldo_final_extrato_mensal = float(saldo_final_extrato_mensal1)
        else:
            saldo_final_extrato_mensal = 0   

    
    remessa_betel_resolução2 = Gerais.objects.values('remessa_Enviada_para_Betel_Resolução').last()
    if remessa_betel_resolução2 is None:
        remessa_betel_resolução = 0   
    else:
        remessa_betel_resolução1 = remessa_betel_resolução2['remessa_Enviada_para_Betel_Resolução']
        if remessa_betel_resolução1:
            remessa_betel_resolução = float(remessa_betel_resolução1)
        else:
            remessa_betel_resolução = 0   

    saldo_final_extrato_betel2 = Gerais.objects.values('saldo_Final_do_Extrato_de_Betel').last()
    if saldo_final_extrato_betel2 is None:
        saldo_final_extrato_betel = 0   
    else:
        saldo_final_extrato_betel1 = saldo_final_extrato_betel2['saldo_Final_do_Extrato_de_Betel']
        if saldo_final_extrato_betel1:
            saldo_final_extrato_betel = float(saldo_final_extrato_betel1)
        else:
            saldo_final_extrato_betel = 0   

    saldo_donativos_anterior2 = Gerais.objects.values('saldo_Final_dos_Donativos_Mês_Anterior').last()
    if saldo_donativos_anterior2 is None:
        saldo_donativos_anterior = 0   
    else:
        saldo_donativos_anterior1 = saldo_donativos_anterior2['saldo_Final_dos_Donativos_Mês_Anterior']
        if saldo_donativos_anterior1:
            saldo_donativos_anterior = float(saldo_donativos_anterior1)
        else:
            saldo_donativos_anterior = 0   

    
    saldo_conta_anterior2 = Gerais.objects.values('saldo_Final_da_Conta_Bancária_Mês_Anterior').last()
    if saldo_conta_anterior2 is None:
        saldo_conta_anterior = 0   
    else:
        saldo_conta_anterior1 = saldo_conta_anterior2['saldo_Final_da_Conta_Bancária_Mês_Anterior']
        if saldo_conta_anterior1:
            saldo_conta_anterior = float(saldo_conta_anterior1)
        else:
            saldo_conta_anterior = 0   

    saldo_betel_anterior2 = Gerais.objects.values('saldo_Final_da_Conta_em_Betel_Mês_Anterior').last()
    if saldo_betel_anterior2 is None:
        saldo_betel_anterior = 0   
    else:
        saldo_betel_anterior1 = saldo_betel_anterior2['saldo_Final_da_Conta_em_Betel_Mês_Anterior']
        if saldo_betel_anterior1:
            saldo_betel_anterior = float(saldo_betel_anterior1)
        else:
            saldo_betel_anterior = 0   
    

    
    total_C10 = Contas.objects.filter(símbolo='C').aggregate(Sum('donativos_Entrada'))
    total_C20 = Contas.objects.filter(símbolo='C').aggregate(Sum('conta_Entrada'))
    total_C30 = Contas.objects.filter(símbolo='C').aggregate(Sum('outra_Entrada'))
    total_Ca = total_C10[('donativos_Entrada__sum')]
    if total_Ca:
        total_C1 = float(total_Ca)
    else:
        total_C1 = 0  
    total_Cb = total_C20[('conta_Entrada__sum')]
    if total_Cb:
        total_C2 = float(total_Cb)
    else:
        total_C2 = 0  
    total_Cc = total_C30[('outra_Entrada__sum')]
    if total_Cc:
        total_C3 = float(total_Cc)
    else:
        total_C3 = 0  

    total_C = total_C1 + total_C2 + total_C3
    
    t_C10 = Contas.objects.filter(símbolo='T').aggregate(Sum('donativos_Entrada'))
    t_C20 = Contas.objects.filter(símbolo='T').aggregate(Sum('conta_Entrada'))
    t_C30 = Contas.objects.filter(símbolo='T').aggregate(Sum('outra_Entrada'))
    t_Ca = t_C10[('donativos_Entrada__sum')]
    if t_Ca:
        t_C1 = float(t_Ca)
    else:
        t_C1 = 0  
    t_Cb = t_C20[('conta_Entrada__sum')]
    if t_Cb:
        t_C2 = float(t_Cb)
    else:
        t_C2 = 0  
    t_Cc = t_C30[('outra_Entrada__sum')]
    if t_Cc:
        t_C3 = float(t_Cc)
    else:
        t_C3 = 0  

    transporte = t_C1 + t_C2 + t_C3

    total_O10 = Contas.objects.filter(símbolo='O').aggregate(Sum('donativos_Entrada'))
    total_O20 = Contas.objects.filter(símbolo='O').aggregate(Sum('conta_Entrada'))
    total_O30 = Contas.objects.filter(símbolo='O').aggregate(Sum('outra_Entrada'))
    total_Oa = total_O10[('donativos_Entrada__sum')]
    if total_Oa:
        total_O1 = float(total_Oa)
    else:
        total_O1 = 0  
    total_Ob = total_O20[('conta_Entrada__sum')]
    if total_Ob:
        total_O2 = float(total_Ob)
    else:
        total_O2 = 0  
    total_Oc = total_O30[('outra_Entrada__sum')]
    if total_Oc:
        total_O3 = float(total_Oc)
    else:
        total_O3 = 0  

    total_O = total_O1 + total_O2 + total_O3

    total_CF10 = Contas.objects.filter(símbolo='CF').aggregate(Sum('donativos_Entrada'))
    total_CF20 = Contas.objects.filter(símbolo='CF').aggregate(Sum('conta_Entrada'))
    total_CF30 = Contas.objects.filter(símbolo='CF').aggregate(Sum('outra_Entrada'))
    total_CFa = total_CF10[('donativos_Entrada__sum')]
    if total_CFa:
        total_CF1 = float(total_CFa)
    else:
        total_CF1 = 0  
    total_CFb = total_CF20[('conta_Entrada__sum')]
    if total_CFb:
        total_CF2 = float(total_CFb)
    else:
        total_CF2 = 0  
    total_CFc = total_CF30[('outra_Entrada__sum')]
    if total_CFc:
        total_CF3 = float(total_CFc)
    else:
        total_CF3 = 0  

    total_CF = total_CF1 + total_CF2 + total_CF3
    
    total_CE10 = Contas.objects.filter(símbolo='CE').aggregate(Sum('donativos_Entrada'))
    total_CE20 = Contas.objects.filter(símbolo='CE').aggregate(Sum('conta_Entrada'))
    total_CE30 = Contas.objects.filter(símbolo='CE').aggregate(Sum('outra_Entrada'))
    total_CEa = total_CE10[('donativos_Entrada__sum')]
    if total_CEa:
        total_CE1 = float(total_CEa)
    else:
        total_CE1 = 0  
    total_CEb = total_CE20[('conta_Entrada__sum')]
    if total_CEb:
        total_CE2 = float(total_CEb)
    else:
        total_CE2 = 0  
    total_CEc = total_CE30[('outra_Entrada__sum')]
    if total_CEc:
        total_CE3 = float(total_CEc)
    else:
        total_CE3 = 0  

    total_CE = total_CE1 + total_CE2 + total_CE3    

    total_C_TE = total_CE + total_C

    total_G10 = Contas.objects.filter(símbolo='G').aggregate(Sum('donativos_Saída'))
    total_G20 = Contas.objects.filter(símbolo='G').aggregate(Sum('conta_Saída'))
    total_G30 = Contas.objects.filter(símbolo='G').aggregate(Sum('outra_Saída'))
    rendimentos3 = Contas.objects.filter(símbolo='J').aggregate(Sum('conta_Entrada'))
    rendimentos_estorno3 = Contas.objects.filter(símbolo='J').aggregate(Sum('conta_Saída'))

    total_Ga = total_G10[('donativos_Saída__sum')]
    if total_Ga:
        total_G1 = float(total_Ga)
    else:
        total_G1 = 0  
    total_Gb = total_G20[('conta_Saída__sum')]
    if total_Gb:
        total_G2 = float(total_Gb)
    else:
        total_G2 = 0  
    total_Gc = total_G30[('outra_Saída__sum')]
    if total_Gc:
        total_G3 = float(total_Gc)
    else:
        total_G3 = 0  

    rendimentos2 = rendimentos3[('conta_Entrada__sum')] 
    if rendimentos2:
        rendimentos_ = float(rendimentos2)
    else:
        rendimentos_ = 0  

    rendimentos_estorno2 = rendimentos_estorno3[('conta_Saída__sum')]
    if rendimentos_estorno2:
        rendimentos_estorno = float(rendimentos_estorno2)
    else: rendimentos_estorno = 0

    rendimentos = rendimentos_ - rendimentos_estorno


    total_G = total_G1 + total_G2 + total_G3
    
    

    outras_ent_1_mot = request.GET.get('outras_ent_1_mot')
    if outras_ent_1_mot is None:
         outras_ent_1_mot = "Outra Finalidade"
         
    outras_ent_1a = request.GET.get('outras_ent_1')
    if outras_ent_1a:
        outras_ent_1 = float(outras_ent_1a)
    else:
        outras_ent_1 = 0
    
    outras_ent_2_mot = request.GET.get('outras_ent_2_mot')
    if outras_ent_2_mot is None:
         outras_ent_2_mot = ""
    outras_ent_2a = request.GET.get('outras_ent_2')
    if outras_ent_2a:
        outras_ent_2 = float(outras_ent_2a)
    else:
        outras_ent_2 = 0  

    outras_sai_1_mot = request.GET.get('outras_sai_1_mot')
    if  outras_sai_1_mot is None:
          outras_sai_1_mot = "Outra Finalidade"
    outras_sai_1a = request.GET.get('outras_sai_1')
    if outras_sai_1a:
        outras_sai_1 = float(outras_sai_1a)
    else:
        outras_sai_1 = 0  
    outras_sai_2_mot = request.GET.get('outras_sai_2_mot')
    if  outras_sai_2_mot is None:
          outras_sai_2_mot = ""
    outras_sai_2a = request.GET.get('outras_sai_2')
    if outras_sai_2a:
        outras_sai_2 = float(outras_sai_2a)
    else:
        outras_sai_2 = 0  

    fundos_reservados1 = request.GET.get('fundos_reservados')
    if fundos_reservados1:
        fundos_reservados = float(fundos_reservados1)
    else:
        fundos_reservados = 0  
        
    fundos_reservados_desc = request.GET.get('fundos_reservados_desc')
    if  fundos_reservados_desc is None:
          fundos_reservados_desc = "Fundos Reservados"    
 
   
    servo1 = Gerais.objects.values('servo_de_contas').last()
    if servo1:
        servo = servo1['servo_de_contas']
        if servo is None:
            servo = ""


    for a in relatório:
        remessa_Enviada_para_Betel_Resolução1 = a.remessa_Enviada_para_Betel_Resolução
        if remessa_Enviada_para_Betel_Resolução1:
            remessa_Enviada_para_Betel_Resolução = float(remessa_Enviada_para_Betel_Resolução1)
        else: 
            remessa_Enviada_para_Betel_Resolução = 0  

    
    total_CE_betel = out_ent_value + total_CE
    betel = total_O + remessa_betel_resolução + total_CF
    sal_final_don = saldo_donativos_anterior + don_ent_value - don_sai_value
    sal_final_con = saldo_conta_anterior + con_ent_value - con_sai_value
    sal_final_betel = saldo_betel_anterior + out_ent_value - out_sai_value
    saldo_conta_extrato_sem_betel = saldo_final_extrato_mensal -betel
    total_total = sal_final_don + sal_final_con + sal_final_betel
    
    saldo_total_mês_anterior = saldo_donativos_anterior + saldo_conta_anterior + saldo_betel_anterior
    
    cong1 = Gerais.objects.values('congregação').last()
    if cong1:
        cong = cong1['congregação']
        if cong is None:
            cong = ""
    mês = ""
    mês1 = Gerais.objects.values('mês').last()
    if mês1:
        mês = mês1['mês']
        if mês is None:
            mês = ""
    ano = ""
    ano1 = Gerais.objects.values('ano').last()   
    if ano1:
        ano = ano1['ano']
        if ano is None:
            ano = "" 

    mês_ano = (f'{mês} / {ano}')
    semi_total_entradas = total_C + rendimentos + out_ent_value + outras_ent_1 + total_CE + transporte
    semi_total_entradas1 = total_O  + outras_ent_2 + total_CF
    total_das_entradas = semi_total_entradas + semi_total_entradas1
    semi_total_saídas = total_G + remessa_betel_resolução + outras_sai_1  
    semi_total_saídas1 = total_O + total_CF + outras_sai_2
    total_das_despesas = semi_total_saídas + semi_total_saídas1
    saldo_mês_atual = saldo_total_mês_anterior + total_das_entradas - total_das_despesas 
    entradas_despesas = total_das_entradas -total_das_despesas 
    saldo_mês_atual_sem_fundos =  saldo_mês_atual - fundos_reservados
    total_betel = total_CF + total_O + remessa_Enviada_para_Betel_Resolução
    

  
    
    if request.method == 'POST':
        
        data_dict = {
                
                "saldo_mês_atual_sem_fundos": saldo_mês_atual_sem_fundos,
                "entradas_despesas": entradas_despesas,
                "semi_total_saídas1": semi_total_saídas1,
                "semi_total_saídas": semi_total_saídas,
                "semi_total_entradas1":semi_total_entradas1,
                "semi_total_entradas":semi_total_entradas,
                "congregação": cong,
                "servo": servo,
                "mês/ano": mês_ano,
                "saldo_total_mês_anterior": saldo_total_mês_anterior,
                "total_C": total_C,
                "out_ent_value": out_ent_value,
                "rendimentos": rendimentos,
                "outras_ent_1_mot":outras_ent_1_mot,
                "outras_ent_2_mot": outras_ent_2_mot,
                "outras_ent_1":outras_ent_1,
                "outras_ent_2":outras_ent_2,
                "total_das_entradas": total_das_entradas,
                "total_G":total_G,
                "total_CF":total_CF,
                "remessa_Enviada_para_Betel_Resolução": remessa_Enviada_para_Betel_Resolução,
                "outras_sai_1_mot": outras_sai_1_mot,
                "outras_sai_2_mot":outras_sai_2_mot,
                "outras_sai_1": outras_sai_1,
                "outras_sai_2": outras_sai_2,
                "total_CF": total_CF,
                "total_O": total_O,
                "total_das_despesas": total_das_despesas,
                "fundos_reservados": fundos_reservados,
                "fundos_reservados_desc": fundos_reservados_desc,
                "saldo_mês_atual": saldo_mês_atual,
                'total_CE':total_CE,
                'transporte':transporte,
                'total_CE_betel':total_CE_betel
               

            }
                

        fillpdfs.write_fillable_pdf('static/relatório.pdf', 'static/relatório_pronto.pdf', data_dict)
        

    context = {
        'saldo_conta_extrato_sem_betel': saldo_conta_extrato_sem_betel,
        'betel': betel,
        'sal_final_con': sal_final_con,
        'saldo_final_extrato_betel': saldo_final_extrato_betel,
        'saldo_donativos_anterior': saldo_donativos_anterior,
        'saldo_conta_anterior': saldo_conta_anterior,
        'saldo_betel_anterior': saldo_betel_anterior,
        'sal_final_don': sal_final_don,
        'sal_final_con': sal_final_con,
        'sal_final_betel': sal_final_betel,
        'total_C': total_C,
        'total_O':total_O,
        'total_G':total_G,
        "total_CF":total_CF,
        'saldo_conta_extrato_sem_betel': saldo_conta_extrato_sem_betel,
        'total_total': total_total,
        'saldo_final_extrato_mensal': saldo_final_extrato_mensal,
        'rendimentos': rendimentos,
        'saldo_total_mês_anterior': saldo_total_mês_anterior,
        'relatório': relatório,
        'out_ent_value': out_ent_value,
        'outras_ent_1': outras_ent_1,
        'outras_ent_2': outras_ent_2,
        'outras_ent_1_mot': outras_ent_1_mot,
        'outras_ent_2_mot': outras_ent_2_mot,
        'outras_sai_1_mot' :outras_sai_1_mot,
        'outras_sai_1': outras_sai_1,
        'outras_sai_2_mot': outras_sai_2_mot,
        'outras_sai_2': outras_sai_2,
        'fundos_reservados':fundos_reservados,
        "fundos_reservados_desc": fundos_reservados_desc,
        'servo': servo,
        'total_das_entradas': total_das_entradas,
        'total_das_despesas': total_das_despesas,
        'saldo_mês_atual': saldo_mês_atual,
        'mês_ano': mês_ano,
        'total_CE':total_CE,
        'transporte':transporte,
        'total_CE_betel':total_CE_betel,
        'total_betel':total_betel,
        
        }
    
    return render(request, 'cp/relatório_mensal.html', context)
   
@login_required
def contas(request):
    global saldo_final_extrato_betel
    global sal_final_betel
    global total_total
    global saldo_conta_extrato_sem_betel 
    global sal_final_con
    global don_ent_value
    global don_sai_value
    global con_ent_value
    global con_sai_value
    global out_ent_value
    global out_sai_value
    global sal_final_don
    global total_C_TE
    global total_O
    global total_G

    
    if request.method == 'POST':
        contas = Form_FContas(request.POST)
        numeros = []
        
        if contas.is_valid():
           contas.save()
           messages.info(request, 'Inserido com sucesso')
           contas = Form_FContas()
            
           return redirect('/contas')
 
    else:
        contas = Form_FContas()
    numeros = Contas.objects.order_by('dia').all()
    don_ent = Contas.objects.aggregate(total=Sum('donativos_Entrada'))
    don_sai = Contas.objects.aggregate(total=Sum('donativos_Saída'))
    con_ent = Contas.objects.aggregate(total=Sum('conta_Entrada'))
    con_sai = Contas.objects.aggregate(total=Sum('conta_Saída'))
    out_ent = Contas.objects.aggregate(total=Sum('outra_Entrada'))
    out_sai = Contas.objects.aggregate(total=Sum('outra_Saída'))
 
    don_ent_value1 = don_ent['total']
    if don_ent_value1:
        don_ent_value = float(don_ent_value1)
    else:
        don_ent_value = 0    
    don_sai_value1 = don_sai['total']
    if don_sai_value1:
        don_sai_value = float(don_sai_value1)
    else:
        don_sai_value = 0    
    con_ent_value1 = con_ent['total']
    if con_ent_value1:
        con_ent_value = float(con_ent_value1)
    else:
        con_ent_value = 0    
    con_sai_value1 = con_sai['total']
    if con_sai_value1:
        con_sai_value = float(con_sai_value1)
    else:
        con_sai_value = 0    
    out_ent_value1 = out_ent['total']
    if out_ent_value1:
        out_ent_value = float(out_ent_value1)
    else:
        out_ent_value = 0    
    out_sai_value1 = out_sai['total']
    if out_sai_value1:
        out_sai_value = float(out_sai_value1)
    else:
        out_sai_value = 0   

    cong = Gerais.objects.values('congregação').last()
    cidade = Gerais.objects.values('cidade').last()
    estado = Gerais.objects.values('estado').last()
    mês = Gerais.objects.values('mês').last()
    data = Gerais.objects.values('data_do_Fechamento').last()
    ano = Gerais.objects.values('ano').last()

    saldo_final_extrato_mensal2 = Gerais.objects.values('saldo_Final_do_Extrato_Mensal').last()
    if saldo_final_extrato_mensal2 is None:
        saldo_final_extrato_mensal = 0   
    else:
        saldo_final_extrato_mensal1 = saldo_final_extrato_mensal2['saldo_Final_do_Extrato_Mensal']
        if saldo_final_extrato_mensal1:
            saldo_final_extrato_mensal = float(saldo_final_extrato_mensal1)
        else:
            saldo_final_extrato_mensal = 0   

  
    remessa_betel_resolução2 = Gerais.objects.values('remessa_Enviada_para_Betel_Resolução').last()
    if remessa_betel_resolução2 is None:
        remessa_betel_resolução = 0   
    else:
        remessa_betel_resolução1 = remessa_betel_resolução2['remessa_Enviada_para_Betel_Resolução']
        if remessa_betel_resolução1:
            remessa_betel_resolução = float(remessa_betel_resolução1)
        else:
            remessa_betel_resolução = 0   

    saldo_final_extrato_betel2 = Gerais.objects.values('saldo_Final_do_Extrato_de_Betel').last()
    if saldo_final_extrato_betel2 is None:
        saldo_final_extrato_betel = 0   
    else:
        saldo_final_extrato_betel1 = saldo_final_extrato_betel2['saldo_Final_do_Extrato_de_Betel']
        if saldo_final_extrato_betel1:
            saldo_final_extrato_betel = float(saldo_final_extrato_betel1)
        else:
            saldo_final_extrato_betel = 0   

    saldo_donativos_anterior2 = Gerais.objects.values('saldo_Final_dos_Donativos_Mês_Anterior').last()
    if saldo_donativos_anterior2 is None:
        saldo_donativos_anterior = 0   
    else:
        saldo_donativos_anterior1 = saldo_donativos_anterior2['saldo_Final_dos_Donativos_Mês_Anterior']
        if saldo_donativos_anterior1:
            saldo_donativos_anterior = float(saldo_donativos_anterior1)
        else:
            saldo_donativos_anterior = 0   

    
    saldo_conta_anterior2 = Gerais.objects.values('saldo_Final_da_Conta_Bancária_Mês_Anterior').last()
    if saldo_conta_anterior2 is None:
        saldo_conta_anterior = 0   
    else:
        saldo_conta_anterior1 = saldo_conta_anterior2['saldo_Final_da_Conta_Bancária_Mês_Anterior']
        if saldo_conta_anterior1:
            saldo_conta_anterior = float(saldo_conta_anterior1)
        else:
            saldo_conta_anterior = 0   

    saldo_betel_anterior2 = Gerais.objects.values('saldo_Final_da_Conta_em_Betel_Mês_Anterior').last()
    if saldo_betel_anterior2 is None:
        saldo_betel_anterior = 0   
    else:
        saldo_betel_anterior1 = saldo_betel_anterior2['saldo_Final_da_Conta_em_Betel_Mês_Anterior']
        if saldo_betel_anterior1:
            saldo_betel_anterior = float(saldo_betel_anterior1)
        else:
            saldo_betel_anterior = 0   
    


    
    total_C10 = Contas.objects.filter(símbolo='C').aggregate(Sum('donativos_Entrada'))
    total_C20 = Contas.objects.filter(símbolo='C').aggregate(Sum('conta_Entrada'))
    total_C30 = Contas.objects.filter(símbolo='C').aggregate(Sum('outra_Entrada'))
    total_Ca = total_C10[('donativos_Entrada__sum')]
    if total_Ca:
        total_C1 = float(total_Ca)
    else:
        total_C1 = 0  
    total_Cb = total_C20[('conta_Entrada__sum')]
    if total_Cb:
        total_C2 = float(total_Cb)
    else:
        total_C2 = 0  
    total_Cc = total_C30[('outra_Entrada__sum')]
    if total_Cc:
        total_C3 = float(total_Cc)
    else:
        total_C3 = 0  

    total_C = total_C1 + total_C2 + total_C3

    total_O10 = Contas.objects.filter(símbolo='O').aggregate(Sum('donativos_Entrada'))
    total_O20 = Contas.objects.filter(símbolo='O').aggregate(Sum('conta_Entrada'))
    total_O30 = Contas.objects.filter(símbolo='O').aggregate(Sum('outra_Entrada'))
    total_Oa = total_O10[('donativos_Entrada__sum')]
    if total_Oa:
        total_O1 = float(total_Oa)
    else:
        total_O1 = 0  
    total_Ob = total_O20[('conta_Entrada__sum')]
    if total_Ob:
        total_O2 = float(total_Ob)
    else:
        total_O2 = 0  
    total_Oc = total_O30[('outra_Entrada__sum')]
    if total_Oc:
        total_O3 = float(total_Oc)
    else:
        total_O3 = 0  

    total_O = total_O1 + total_O2 + total_O3

    total_G10 = Contas.objects.filter(símbolo='G').aggregate(Sum('donativos_Saída'))
    total_G20 = Contas.objects.filter(símbolo='G').aggregate(Sum('conta_Saída'))
    total_G30 = Contas.objects.filter(símbolo='G').aggregate(Sum('outra_Saída'))
    

    total_Ga = total_G10[('donativos_Saída__sum')]
    if total_Ga:
        total_G1 = float(total_Ga)
    else:
        total_G1 = 0  
    total_Gb = total_G20[('conta_Saída__sum')]
    if total_Gb:
        total_G2 = float(total_Gb)
    else:
        total_G2 = 0  
    total_Gc = total_G30[('outra_Saída__sum')]
    if total_Gc:
        total_G3 = float(total_Gc)
    else:
        total_G3 = 0  

    total_CF10 = Contas.objects.filter(símbolo='CF').aggregate(Sum('donativos_Entrada'))
    total_CF20 = Contas.objects.filter(símbolo='CF').aggregate(Sum('conta_Entrada'))
    total_CF30 = Contas.objects.filter(símbolo='CF').aggregate(Sum('outra_Entrada'))
    total_CFa = total_CF10[('donativos_Entrada__sum')]
    if total_CFa:
        total_CF1 = float(total_CFa)
    else:
        total_CF1 = 0  
    total_CFb = total_CF20[('conta_Entrada__sum')]
    if total_CFb:
        total_CF2 = float(total_CFb)
    else:
        total_CF2 = 0  
    total_CFc = total_CF30[('outra_Entrada__sum')]
    if total_CFc:
        total_CF3 = float(total_CFc)
    else:
        total_CF3 = 0  

    total_CF = total_CF1 + total_CF2 + total_CF3    

   

    total_G = total_G1 + total_G2 + total_G3
    
    total_CE10 = Contas.objects.filter(símbolo='CE').aggregate(Sum('donativos_Entrada'))
    total_CE20 = Contas.objects.filter(símbolo='CE').aggregate(Sum('conta_Entrada'))
    total_CE30 = Contas.objects.filter(símbolo='CE').aggregate(Sum('outra_Entrada'))
    total_CEa = total_CE10[('donativos_Entrada__sum')]
    if total_CEa:
        total_CE1 = float(total_CEa)
    else:
        total_CE1 = 0  
    total_CEb = total_CE20[('conta_Entrada__sum')]
    if total_CEb:
        total_CE2 = float(total_CEb)
    else:
        total_CE2 = 0  
    total_CEc = total_CE30[('outra_Entrada__sum')]
    if total_CEc:
        total_CE3 = float(total_CEc)
    else:
        total_CE3 = 0  

    total_CE = total_CE1 + total_CE2 + total_CE3    

    total_C_TE = total_CE + total_C
    betel = total_O + remessa_betel_resolução + total_CF
    sal_final_don = saldo_donativos_anterior + don_ent_value - don_sai_value
    sal_final_con = saldo_conta_anterior + con_ent_value - con_sai_value
    sal_final_betel = saldo_betel_anterior + out_ent_value - out_sai_value
    saldo_conta_extrato_sem_betel = saldo_final_extrato_mensal - betel
    total_total = sal_final_don + sal_final_con + sal_final_betel
    saldo_total_mês_anterior = saldo_donativos_anterior + saldo_conta_anterior + saldo_betel_anterior
    
    
    context = {
        'numeros': numeros,
        'contas': contas,
        'don_ent_value': don_ent_value,
        'don_sai_value': don_sai_value,
        'con_ent_value': con_ent_value,
        'con_sai_value': con_sai_value,
        'out_ent_value': out_ent_value,
        'out_sai_value': out_sai_value,
        'mês': mês,
        'cong': cong,
        'data': data,
        'cidade': cidade,
        'estado': estado,
        'ano': ano,
        'saldo_conta_extrato_sem_betel': saldo_conta_extrato_sem_betel,
        'betel': betel,
        'sal_final_con': sal_final_con,
        'saldo_final_extrato_betel': saldo_final_extrato_betel,
        'saldo_donativos_anterior': saldo_donativos_anterior,
        'saldo_conta_anterior': saldo_conta_anterior,
        'saldo_betel_anterior': saldo_betel_anterior,
        'sal_final_don': sal_final_don,
        'sal_final_con': sal_final_con,
        'sal_final_betel': sal_final_betel,
        'total_C_TE': total_C_TE,
        'total_O':total_O,
        'total_G':total_G,
        'saldo_conta_extrato_sem_betel': saldo_conta_extrato_sem_betel,
        'total_total': total_total,
        'saldo_final_extrato_mensal': saldo_final_extrato_mensal,
        'saldo_total_mês_anterior': saldo_total_mês_anterior,
        'saldo_mês_atual': saldo_mês_atual,
       
    }
    
    return render(request, 'cp/folha.html', context )  

@login_required
def registro (request):
    relatório = Gerais.objects.all()

    total_CF10 = Contas.objects.filter(símbolo='CF').aggregate(Sum('donativos_Entrada'))
    total_CF20 = Contas.objects.filter(símbolo='CF').aggregate(Sum('conta_Entrada'))
    total_CF30 = Contas.objects.filter(símbolo='CF').aggregate(Sum('outra_Entrada'))
    total_CFa = total_CF10[('donativos_Entrada__sum')]
    if total_CFa:
        total_CF1 = float(total_CFa)
    else:
        total_CF1 = 0  
    total_CFb = total_CF20[('conta_Entrada__sum')]
    if total_CFb:
        total_CF2 = float(total_CFb)
    else:
        total_CF2 = 0  
    total_CFc = total_CF30[('outra_Entrada__sum')]
    if total_CFc:
        total_CF3 = float(total_CFc)
    else:
        total_CF3 = 0  

    total_CF = total_CF1 + total_CF2 + total_CF3
    
    total_O10 = Contas.objects.filter(símbolo='O').aggregate(Sum('donativos_Entrada'))
    total_O20 = Contas.objects.filter(símbolo='O').aggregate(Sum('conta_Entrada'))
    total_O30 = Contas.objects.filter(símbolo='O').aggregate(Sum('outra_Entrada'))
    total_Oa = total_O10[('donativos_Entrada__sum')]
    if total_Oa:
        total_O1 = float(total_Oa)
    else:
        total_O1 = 0  
    total_Ob = total_O20[('conta_Entrada__sum')]
    if total_Ob:
        total_O2 = float(total_Ob)
    else:
        total_O2 = 0  
    total_Oc = total_O30[('outra_Entrada__sum')]
    if total_Oc:
        total_O3 = float(total_Oc)
    else:
        total_O3 = 0  

    total_O = total_O1 + total_O2 + total_O3
    
    
    
    remessa_betel_resolução2 = Gerais.objects.values('remessa_Enviada_para_Betel_Resolução').last()
    if remessa_betel_resolução2 is None:
        remessa_betel_resolução = 0   
    else:
        remessa_betel_resolução1 = remessa_betel_resolução2['remessa_Enviada_para_Betel_Resolução']
        if remessa_betel_resolução1:
            remessa_betel_resolução = float(remessa_betel_resolução1)
        else:
            remessa_betel_resolução = 0   

    confirmação = request.GET.get('confirmação')
    
    preenchido1 = Gerais.objects.values('servo_de_contas').last()
    if preenchido1:
        preenchido = preenchido1['servo_de_contas']
        if preenchido is None:
            preenchido = ""
    
    aprovado1 = Gerais.objects.values('secretário').last()
    if aprovado1:
        aprovado = aprovado1['secretário']
        if aprovado is None:
            aprovado = ""
    
    
    
    finalidade1 = request.GET.get('finalidades1')
    finalidade2 = request.GET.get('finalidades2') 
    valor_da_finalidade1a = request.GET.get('valor_finalidades1')
    if valor_da_finalidade1a:
        valor_da_finalidade1 = float(valor_da_finalidade1a)
    else:
        valor_da_finalidade1 = 0  
    valor_da_finalidade2a = request.GET.get('valor_finalidades2')
    if valor_da_finalidade2a:
        valor_da_finalidade2 = float(valor_da_finalidade2a)
    else:
        valor_da_finalidade2 = 0  

    for a in relatório:
        cong = a.congregação
   

    for a in relatório:
        remessa_Enviada_para_Betel_Resolução1 = a.remessa_Enviada_para_Betel_Resolução
        if remessa_Enviada_para_Betel_Resolução1:
            remessa_Enviada_para_Betel_Resolução = float(remessa_Enviada_para_Betel_Resolução1)
        else: 
            remessa_Enviada_para_Betel_Resolução = 0   

        

    total_enviados = valor_da_finalidade1 + valor_da_finalidade2 + total_O + remessa_betel_resolução + total_CF
    data_da_transação = request.GET.get('data')
    
    
    if request.method == 'POST':
        data_da_transação2 = datetime.strptime(data_da_transação,"%Y-%m-%d" )
        data_da_transação = data_da_transação2.strftime("%d-%m-%Y")
        
        

        data_dict = {
                "total_CF": total_CF,
                "cong": cong,
                "remessa_betel_resolução": remessa_betel_resolução,
                "total_O": total_O,
                'data_da_transação':data_da_transação,
                'confirmação': confirmação,
                'aprovado': aprovado,
                'preenchido':preenchido,
                'finalidade1': finalidade1,
                'finalidade2': finalidade2,
                'valor_da_finalidade2': valor_da_finalidade2,
                'valor_da_finalidade1':valor_da_finalidade1,
                'total_enviados':total_enviados,

            }
                
        fillpdfs.write_fillable_pdf('static/registro.pdf', 'static/registro_pronto.pdf', data_dict)
    
    context = {
        'relatório': relatório,
        'data_da_transação':data_da_transação,
        'confirmação': confirmação,
        'aprovado': aprovado,
        'preenchido':preenchido,
        'finalidade1': finalidade1,
        'finalidade2': finalidade2,
        'valor_da_finalidade2': valor_da_finalidade2,
        'valor_da_finalidade1':valor_da_finalidade1,
        'total_O':total_O,
        'remessa_betel_resolução': remessa_betel_resolução,
        'total_enviados':total_enviados,
        "total_CF": total_CF

    }
    
    return render(request, 'cp/registro.html', context)

@login_required
def recibo(request, id):

    contas = get_object_or_404(Contas, pk=id)
    form = Form_FContas(instance=contas)
 
    if request.method == 'POST':
        form = Form_FContas(request.POST, instance=contas)
         
        if form.is_valid():
            contas.save()
            messages.info(request, 'Editado com sucesso')
            return redirect('/contas')
   
        else:
            return render(request, 'cp/recibo.html', {'form':form ,'contas': contas})  


    return render(request, 'cp/recibo.html', {'form':form ,'contas': contas})  

@login_required
def imprimir_FC(request):
    if request.method == 'POST':
        cong1 = Gerais.objects.values('congregação').last()
        cong = cong1['congregação']
        if cong is None:
            cong = ""
        cidade1 = Gerais.objects.values('cidade').last()
        cidade = cidade1['cidade']
        if cidade is None:
            cidade = ""
        estado1 = Gerais.objects.values('estado').last()
        estado = estado1['estado']
        if estado is None:
            estado = ""
        mês1 = Gerais.objects.values('mês').last()
        mês = mês1['mês']
        if mês is None:
            mês = ""
        data2 = Gerais.objects.values('data_do_Fechamento').last()
        data1 = data2['data_do_Fechamento']
        data = data1.strftime("%d-%m-%Y")
        if data is None:
           data = ""
        ano1 = Gerais.objects.values('ano').last()
        ano = ano1['ano']
        if ano is None:
            ano = ""
        números = Contas.objects.order_by('dia').all().values_list()
        df = pd.DataFrame(números)
        
        designacao = load_workbook("static/FC.xlsx")
        writer = pd.ExcelWriter("static/FC_preenchido.xlsx")
        writer.book = designacao
        writer.sheets = dict((ws.title, ws) for ws in designacao.worksheets)
        df.to_excel(writer, 'HOME', startrow=0, startcol=0, header=False, index=False)
        writer.save()

        df_preenchido = openpyxl.load_workbook("static/FC_preenchido.xlsx")
        df_trabalho = df_preenchido.active    
        cel1 = df_trabalho["B1"]
        cel2 = df_trabalho["C1"]
        cel3 = df_trabalho["D1"]
        cel4 = df_trabalho["E1"]
        cel5 = df_trabalho["F1"]
        cel6 = df_trabalho["G1"]
        cel7 = df_trabalho["H1"]
        cel8 = df_trabalho["I1"]
        cel9 = df_trabalho["J1"]
        cel10 = df_trabalho["B2"]
        cel11 = df_trabalho["C2"]
        cel12 = df_trabalho["D2"]
        cel13 = df_trabalho["E2"]
        cel14 = df_trabalho["F2"]
        cel15 = df_trabalho["G2"]
        cel16 = df_trabalho["H2"]
        cel17 = df_trabalho["I2"]
        cel18 = df_trabalho["J2"]
        cel19 = df_trabalho["B3"]
        cel20 = df_trabalho["C3"]
        cel21 = df_trabalho["D3"]
        cel22 = df_trabalho["E3"]
        cel23 = df_trabalho["F3"]
        cel24 = df_trabalho["G3"]
        cel25 = df_trabalho["H3"]
        cel26 = df_trabalho["I3"]
        cel27 = df_trabalho["J3"]
        cel28 = df_trabalho["B4"]
        cel29 = df_trabalho["C4"]
        cel30 = df_trabalho["D4"]
        cel31 = df_trabalho["E4"]
        cel32 = df_trabalho["F4"]
        cel33 = df_trabalho["G4"]
        cel34 = df_trabalho["H4"]
        cel35 = df_trabalho["I4"]
        cel36 = df_trabalho["J4"]
        cel37 = df_trabalho["B5"]
        cel38 = df_trabalho["C5"]
        cel39 = df_trabalho["D5"]
        cel40 = df_trabalho["E5"]
        cel41 = df_trabalho["F5"]
        cel42 = df_trabalho["G5"]
        cel43 = df_trabalho["H5"]
        cel44 = df_trabalho["I5"]
        cel45 = df_trabalho["J5"]
        cel46 = df_trabalho["B6"]
        cel47 = df_trabalho["C6"]
        cel48 = df_trabalho["D6"]
        cel49 = df_trabalho["E6"]
        cel50 = df_trabalho["F6"]
        cel51 = df_trabalho["G6"]
        cel52 = df_trabalho["H6"]
        cel53 = df_trabalho["I6"]
        cel54 = df_trabalho["J6"]
        cel55 = df_trabalho["B7"]
        cel56 = df_trabalho["C7"]
        cel57 = df_trabalho["D7"]
        cel58 = df_trabalho["E7"]
        cel59 = df_trabalho["F7"]
        cel60 = df_trabalho["G7"]
        cel61 = df_trabalho["H7"]
        cel62 = df_trabalho["I7"]
        cel63 = df_trabalho["J7"]
        cel64 = df_trabalho["B8"]
        cel65 = df_trabalho["C8"]
        cel66 = df_trabalho["D8"]
        cel67 = df_trabalho["E8"]
        cel68 = df_trabalho["F8"]
        cel69 = df_trabalho["G8"]
        cel70 = df_trabalho["H8"]
        cel71 = df_trabalho["I8"]
        cel72 = df_trabalho["J8"]
        cel73 = df_trabalho["B9"]
        cel74 = df_trabalho["C9"]
        cel75 = df_trabalho["D9"]
        cel76 = df_trabalho["E9"]
        cel77 = df_trabalho["F9"]
        cel78 = df_trabalho["G9"]
        cel79 = df_trabalho["H9"]
        cel80 = df_trabalho["I9"]
        cel81 = df_trabalho["J9"]
        cel82 = df_trabalho["B10"]
        cel83 = df_trabalho["C10"]
        cel84 = df_trabalho["D10"]
        cel85 = df_trabalho["E10"]
        cel86 = df_trabalho["F10"]
        cel87 = df_trabalho["G10"]
        cel88 = df_trabalho["H10"]
        cel89 = df_trabalho["I10"]
        cel90 = df_trabalho["J10"]
        cel91 = df_trabalho["B11"]
        cel92 = df_trabalho["C11"]
        cel93 = df_trabalho["D11"]
        cel94 = df_trabalho["E11"]
        cel95 = df_trabalho["F11"]
        cel96 = df_trabalho["G11"]
        cel97 = df_trabalho["H11"]
        cel98 = df_trabalho["I11"]
        cel99 = df_trabalho["J11"]
        cel100 = df_trabalho["B12"]
        cel101 = df_trabalho["C12"]
        cel102 = df_trabalho["D12"]
        cel103 = df_trabalho["E12"]
        cel104 = df_trabalho["F12"]
        cel105 = df_trabalho["G12"]
        cel106 = df_trabalho["H12"]
        cel107 = df_trabalho["I12"]
        cel108 = df_trabalho["J12"]
        cel109 = df_trabalho["B13"]
        cel110 = df_trabalho["C13"]
        cel111 = df_trabalho["D13"]
        cel112 = df_trabalho["E13"]
        cel113 = df_trabalho["F13"]
        cel114 = df_trabalho["G13"]
        cel115 = df_trabalho["H13"]
        cel116 = df_trabalho["I13"]
        cel117 = df_trabalho["J13"]
        cel118 = df_trabalho["B14"]
        cel119 = df_trabalho["C15"]
        cel120 = df_trabalho["D15"]
        cel121 = df_trabalho["E15"]
        cel122 = df_trabalho["F15"]
        cel123 = df_trabalho["G15"]
        cel124 = df_trabalho["H15"]
        cel125 = df_trabalho["I15"]
        cel126 = df_trabalho["J15"]
        cel127 = df_trabalho["B16"]
        cel128 = df_trabalho["C16"]
        cel129 = df_trabalho["D16"]
        cel130 = df_trabalho["E16"]
        cel131 = df_trabalho["F16"]
        cel132 = df_trabalho["G16"]
        cel133 = df_trabalho["H16"]
        cel134 = df_trabalho["I16"]
        cel135 = df_trabalho["J16"]
        cel136 = df_trabalho["B17"]
        cel137 = df_trabalho["C17"]
        cel138 = df_trabalho["D17"]
        cel139 = df_trabalho["E17"]
        cel140 = df_trabalho["F17"]
        cel141 = df_trabalho["G17"]
        cel142 = df_trabalho["H17"]
        cel143 = df_trabalho["I17"]
        cel144 = df_trabalho["J17"]
        cel145 = df_trabalho["B18"]
        cel146 = df_trabalho["C18"]
        cel147 = df_trabalho["D18"]
        cel148 = df_trabalho["E18"]
        cel149 = df_trabalho["F18"]
        cel150 = df_trabalho["G18"]
        cel151 = df_trabalho["H18"]
        cel152 = df_trabalho["I18"]
        cel153 = df_trabalho["J18"]
        cel154 = df_trabalho["B19"]
        cel155 = df_trabalho["C19"]
        cel156 = df_trabalho["D19"]
        cel157 = df_trabalho["E19"]
        cel158 = df_trabalho["F19"]
        cel159 = df_trabalho["G19"]
        cel160 = df_trabalho["H19"]
        cel161 = df_trabalho["I19"]
        cel162 = df_trabalho["J19"]
        cel163 = df_trabalho["B20"]
        cel164 = df_trabalho["C20"]
        cel165 = df_trabalho["D20"]
        cel166 = df_trabalho["E20"]
        cel167 = df_trabalho["F20"]
        cel168 = df_trabalho["G20"]
        cel169 = df_trabalho["H20"]
        cel170 = df_trabalho["I20"]
        cel171 = df_trabalho["J20"]
        cel172 = df_trabalho["B21"]
        cel173 = df_trabalho["C21"]
        cel174 = df_trabalho["D21"]
        cel175 = df_trabalho["E21"]
        cel176 = df_trabalho["F21"]
        cel177 = df_trabalho["G21"]
        cel178 = df_trabalho["H21"]
        cel179 = df_trabalho["I21"]
        cel180 = df_trabalho["J21"]
        cel181 = df_trabalho["B22"]
        cel182 = df_trabalho["C22"]
        cel183 = df_trabalho["D22"]
        cel184 = df_trabalho["E22"]
        cel185 = df_trabalho["F22"]
        cel186 = df_trabalho["G22"]
        cel187 = df_trabalho["H22"]
        cel188 = df_trabalho["I22"]
        cel189 = df_trabalho["J22"]
        cel190 = df_trabalho["B23"]
        cel191 = df_trabalho["C23"]
        cel192 = df_trabalho["D23"]
        cel193 = df_trabalho["E23"]
        cel194 = df_trabalho["F23"]
        cel195 = df_trabalho["G23"]
        cel196 = df_trabalho["H23"]
        cel197 = df_trabalho["I23"]
        cel198 = df_trabalho["J23"]
        cel199 = df_trabalho["B24"]
        cel200 = df_trabalho["C24"]
        cel201 = df_trabalho["D24"]
        cel202 = df_trabalho["E24"]
        cel203 = df_trabalho["F24"]
        cel204 = df_trabalho["G24"]
        cel205 = df_trabalho["H24"]
        cel206 = df_trabalho["I24"]
        cel207 = df_trabalho["J24"]
        cel208 = df_trabalho["B25"]
        cel209 = df_trabalho["C25"]
        cel210 = df_trabalho["D25"]
        cel211 = df_trabalho["E25"]
        cel212 = df_trabalho["F25"]
        cel213 = df_trabalho["G25"]
        cel214 = df_trabalho["H25"]
        cel215 = df_trabalho["I25"]
        cel216 = df_trabalho["J25"]
        cel217 = df_trabalho["B26"]
        cel218 = df_trabalho["C26"]
        cel219 = df_trabalho["D26"]
        cel220 = df_trabalho["E26"]
        cel221 = df_trabalho["F26"]
        cel222 = df_trabalho["G26"]
        cel223 = df_trabalho["H26"]
        cel224 = df_trabalho["I26"]
        cel225 = df_trabalho["J26"]
        cel226 = df_trabalho["B27"]
        cel227 = df_trabalho["C27"]
        cel228 = df_trabalho["D27"]
        cel229 = df_trabalho["E27"]
        cel230 = df_trabalho["F27"]
        cel231 = df_trabalho["G27"]
        cel232 = df_trabalho["H27"]
        cel233 = df_trabalho["I27"]
        cel234 = df_trabalho["J27"]
        cel235 = df_trabalho["B28"]
        cel236 = df_trabalho["C28"]
        cel237 = df_trabalho["D28"]
        cel238 = df_trabalho["E28"]
        cel239 = df_trabalho["F28"]
        cel240 = df_trabalho["G28"]
        cel241 = df_trabalho["H28"]
        cel242 = df_trabalho["I28"]
        cel243 = df_trabalho["J28"]
        cel244 = df_trabalho["B29"]
        cel245 = df_trabalho["C29"]
        cel246 = df_trabalho["D29"]
        cel247 = df_trabalho["E29"]
        cel248 = df_trabalho["F29"]
        cel249 = df_trabalho["G29"]
        cel250 = df_trabalho["H29"]
        cel251 = df_trabalho["I29"]
        cel252 = df_trabalho["J29"]
        cel253 = df_trabalho["B30"]
        cel254 = df_trabalho["C30"]
        cel255 = df_trabalho["D30"]
        cel256 = df_trabalho["E30"]
        cel257 = df_trabalho["F30"]
        cel258 = df_trabalho["G30"]
        cel259 = df_trabalho["H30"]
        cel260 = df_trabalho["I30"]
        cel261 = df_trabalho["J30"]
        cel262 = df_trabalho["B31"]
        cel263 = df_trabalho["C31"]
        cel264 = df_trabalho["D31"]
        cel265 = df_trabalho["E31"]
        cel266 = df_trabalho["F31"]
        cel267 = df_trabalho["G31"]
        cel268 = df_trabalho["H31"]
        cel269 = df_trabalho["I31"]
        cel270 = df_trabalho["J31"]
        cel271 = df_trabalho["B32"]
        cel272 = df_trabalho["C32"]
        cel273 = df_trabalho["D32"]
        cel274 = df_trabalho["E32"]
        cel275 = df_trabalho["F32"]
        cel276 = df_trabalho["G32"]
        cel277 = df_trabalho["H32"]
        cel278 = df_trabalho["I32"]
        cel279 = df_trabalho["J32"]
        cel280 = df_trabalho["B33"]
        cel281 = df_trabalho["C33"]
        cel282 = df_trabalho["D33"]
        cel283 = df_trabalho["E33"]
        cel284 = df_trabalho["F33"]
        cel285 = df_trabalho["G33"]
        cel286 = df_trabalho["H33"]
        cel287 = df_trabalho["I33"]
        cel288 = df_trabalho["J33"]
        cel289 = df_trabalho["B34"]
        cel290 = df_trabalho["C34"]
        cel291 = df_trabalho["D34"]
        cel292 = df_trabalho["E34"]
        cel293 = df_trabalho["F34"]
        cel294 = df_trabalho["G34"]
        cel295 = df_trabalho["H34"]
        cel296 = df_trabalho["I34"]
        cel297 = df_trabalho["J34"]
        cel298 = df_trabalho["B35"]
        cel299 = df_trabalho["C35"]
        cel300 = df_trabalho["D35"]
        cel301 = df_trabalho["E35"]
        cel302 = df_trabalho["F35"]
        cel303 = df_trabalho["G35"]
        cel304 = df_trabalho["H35"]
        cel305 = df_trabalho["I35"]
        cel306 = df_trabalho["J35"]
        cel307 = df_trabalho["B36"]
        cel308 = df_trabalho["C36"]
        cel309 = df_trabalho["D36"]
        cel310 = df_trabalho["E36"]
        cel311 = df_trabalho["F36"]
        cel312 = df_trabalho["G36"]
        cel313 = df_trabalho["H36"]
        cel314 = df_trabalho["I36"]
        cel315 = df_trabalho["J36"]
        cel316 = df_trabalho["B37"]
        cel317 = df_trabalho["C37"]
        cel318 = df_trabalho["D37"]
        cel319 = df_trabalho["E37"]
        cel320 = df_trabalho["F37"]
        cel321 = df_trabalho["G37"]
        cel322 = df_trabalho["H37"]
        cel323 = df_trabalho["I37"]
        cel324 = df_trabalho["J37"]
        cel325 = df_trabalho["B38"]
        cel326 = df_trabalho["C38"]
        cel327 = df_trabalho["D38"]
        cel328 = df_trabalho["E38"]
        cel329 = df_trabalho["F38"]
        cel330 = df_trabalho["G38"]
        cel331 = df_trabalho["H38"]
        cel332 = df_trabalho["I38"]
        cel333 = df_trabalho["J38"]
        cel334 = df_trabalho["B39"]
        cel335 = df_trabalho["C39"]
        cel336 = df_trabalho["D39"]
        cel337 = df_trabalho["E39"]
        cel338 = df_trabalho["F39"]
        cel339 = df_trabalho["G39"]
        cel340 = df_trabalho["H39"]
        cel341 = df_trabalho["I39"]
        cel342 = df_trabalho["J39"]
        cel343 = df_trabalho["B40"]
        cel344 = df_trabalho["C40"]
        cel345 = df_trabalho["D40"]
        cel346 = df_trabalho["E40"]
        cel347 = df_trabalho["F40"]
        cel348 = df_trabalho["G40"]
        cel349 = df_trabalho["H40"]
        cel350 = df_trabalho["I40"]
        cel351 = df_trabalho["J40"]
        cel352 = df_trabalho["B41"]
        cel353 = df_trabalho["C41"]
        cel354 = df_trabalho["D41"]
        cel355 = df_trabalho["E41"]
        cel356 = df_trabalho["F41"]
        cel357 = df_trabalho["G41"]
        cel358 = df_trabalho["H41"]
        cel359 = df_trabalho["I41"]
        cel360 = df_trabalho["J41"]
        cel361 = df_trabalho["B42"]
        cel362 = df_trabalho["C42"]
        cel363 = df_trabalho["D42"]
        cel364 = df_trabalho["E42"]
        cel365 = df_trabalho["F42"]
        cel366 = df_trabalho["G42"]
        cel367 = df_trabalho["H42"]
        cel368 = df_trabalho["I42"]
        cel369 = df_trabalho["J42"]
        cel370 = df_trabalho["B43"]
        cel371 = df_trabalho["C43"]
        cel372 = df_trabalho["D43"]
        cel373 = df_trabalho["E43"]
        cel374 = df_trabalho["F43"]
        cel375 = df_trabalho["G43"]
        cel376 = df_trabalho["H43"]
        cel377 = df_trabalho["I43"]
        cel378 = df_trabalho["J43"]
        cel379 = df_trabalho["B44"]
        cel380 = df_trabalho["C44"]
        cel381 = df_trabalho["D44"]
        cel382 = df_trabalho["E44"]
        cel383 = df_trabalho["F44"]
        cel384 = df_trabalho["G44"]
        cel385 = df_trabalho["H44"]
        cel386 = df_trabalho["I44"]
        cel387 = df_trabalho["J44"]
        cel388 = df_trabalho["B45"]
        cel389 = df_trabalho["C45"]
        cel390 = df_trabalho["D45"]
        cel391 = df_trabalho["E45"]
        cel392 = df_trabalho["F45"]
        cel393 = df_trabalho["G45"]
        cel394 = df_trabalho["H45"]
        cel395 = df_trabalho["I45"]
        cel396 = df_trabalho["J45"]
        cel397 = df_trabalho["B46"]
        cel398 = df_trabalho["C46"]
        cel399 = df_trabalho["D46"]
        cel400 = df_trabalho["E46"]
        cel401 = df_trabalho["F46"]
        cel402 = df_trabalho["G46"]
        cel403 = df_trabalho["H46"]
        cel404 = df_trabalho["I46"]
        cel405 = df_trabalho["J46"]
        cel406 = df_trabalho["B47"]
        cel407 = df_trabalho["C47"]
        cel408 = df_trabalho["D47"]
        cel409 = df_trabalho["E47"]
        cel410 = df_trabalho["F47"]
        cel411 = df_trabalho["G47"]
        cel412 = df_trabalho["H47"]
        cel413 = df_trabalho["I47"]
        cel414 = df_trabalho["J47"]
        cel415 = df_trabalho["B48"]
        cel416 = df_trabalho["C48"]
        cel417 = df_trabalho["D48"]
        cel418 = df_trabalho["E48"]
        cel419 = df_trabalho["F48"]
        cel420 = df_trabalho["G48"]
        cel421 = df_trabalho["H48"]
        cel422 = df_trabalho["I48"]
        cel423 = df_trabalho["J48"]
        cel424 = df_trabalho["B49"]
        cel425 = df_trabalho["C49"]
        cel426 = df_trabalho["D49"]
        cel427 = df_trabalho["E49"]
        cel428 = df_trabalho["F49"]
        cel429 = df_trabalho["G49"]
        cel430 = df_trabalho["H49"]
        cel431 = df_trabalho["I49"]
        cel432 = df_trabalho["J49"]
        cel433 = df_trabalho["B49"]
        cel434 = df_trabalho["C49"]
        cel435 = df_trabalho["D49"]
        cel436 = df_trabalho["E49"]
        cel437 = df_trabalho["F49"]
        cel438 = df_trabalho["G49"]
        cel439 = df_trabalho["H49"]
        cel440 = df_trabalho["I49"]
        cel441 = df_trabalho["J49"]
        cel442 = df_trabalho["B50"]
        cel443 = df_trabalho["C50"]
        cel444 = df_trabalho["D50"]
        cel445 = df_trabalho["E50"]
        cel446 = df_trabalho["F50"]
        cel447 = df_trabalho["G50"]
        cel448 = df_trabalho["H50"]
        cel449 = df_trabalho["I50"]
        cel450 = df_trabalho["J50"]
        cel451 = df_trabalho["B51"]
        cel452 = df_trabalho["C51"]
        cel453 = df_trabalho["D51"]
        cel454 = df_trabalho["E51"]
        cel455 = df_trabalho["F51"]
        cel456 = df_trabalho["G51"]
        cel457 = df_trabalho["H51"]
        cel458 = df_trabalho["I51"]
        cel459 = df_trabalho["J51"]
        cel460 = df_trabalho["B52"]
        cel461 = df_trabalho["C52"]
        cel462 = df_trabalho["D52"]
        cel463 = df_trabalho["E52"]
        cel464 = df_trabalho["F52"]
        cel465 = df_trabalho["G52"]
        cel466 = df_trabalho["H52"]
        cel467 = df_trabalho["I52"]
        cel468 = df_trabalho["J52"]
       

        if cel1 .value is None: cel1 .value = ""
        if cel2 .value is None: cel2 .value = ""
        if cel3 .value is None: cel3 .value = ""
        if cel4 .value is None: cel4 .value = ""
        if cel5 .value is None: cel5 .value = ""
        if cel6 .value is None: cel6 .value = ""
        if cel7 .value is None: cel7 .value = ""
        if cel8 .value is None: cel8 .value = ""
        if cel9 .value is None: cel9 .value = ""
        if cel10 .value is None: cel10 .value = ""
        if cel11 .value is None: cel11 .value = ""
        if cel12 .value is None: cel12 .value = ""
        if cel13 .value is None: cel13 .value = ""
        if cel14 .value is None: cel14 .value = ""
        if cel15 .value is None: cel15 .value = ""
        if cel16 .value is None: cel16 .value = ""
        if cel17 .value is None: cel17 .value = ""
        if cel18 .value is None: cel18 .value = ""
        if cel19 .value is None: cel19 .value = ""
        if cel20 .value is None: cel20 .value = ""
        if cel21 .value is None: cel21 .value = ""
        if cel22 .value is None: cel22 .value = ""
        if cel23 .value is None: cel23 .value = ""
        if cel24 .value is None: cel24 .value = ""
        if cel25 .value is None: cel25 .value = ""
        if cel26 .value is None: cel26 .value = ""
        if cel27 .value is None: cel27 .value = ""
        if cel28 .value is None: cel28 .value = ""
        if cel29 .value is None: cel29 .value = ""
        if cel30 .value is None: cel30 .value = ""
        if cel31 .value is None: cel31 .value = ""
        if cel32 .value is None: cel32 .value = ""
        if cel33 .value is None: cel33 .value = ""
        if cel34 .value is None: cel34 .value = ""
        if cel35 .value is None: cel35 .value = ""
        if cel36 .value is None: cel36 .value = ""
        if cel37 .value is None: cel37 .value = ""
        if cel38 .value is None: cel38 .value = ""
        if cel39 .value is None: cel39 .value = ""
        if cel40 .value is None: cel40 .value = ""
        if cel41 .value is None: cel41 .value = ""
        if cel42 .value is None: cel42 .value = ""
        if cel43 .value is None: cel43 .value = ""
        if cel44 .value is None: cel44 .value = ""
        if cel45 .value is None: cel45 .value = ""
        if cel46 .value is None: cel46 .value = ""
        if cel47 .value is None: cel47 .value = ""
        if cel48 .value is None: cel48 .value = ""
        if cel49 .value is None: cel49 .value = ""
        if cel50 .value is None: cel50 .value = ""
        if cel51 .value is None: cel51 .value = ""
        if cel52 .value is None: cel52 .value = ""
        if cel53 .value is None: cel53 .value = ""
        if cel54 .value is None: cel54 .value = ""
        if cel55 .value is None: cel55 .value = ""
        if cel56 .value is None: cel56 .value = ""
        if cel57 .value is None: cel57 .value = ""
        if cel58 .value is None: cel58 .value = ""
        if cel59 .value is None: cel59 .value = ""
        if cel60 .value is None: cel60 .value = ""
        if cel61 .value is None: cel61 .value = ""
        if cel62 .value is None: cel62 .value = ""
        if cel63 .value is None: cel63 .value = ""
        if cel64 .value is None: cel64 .value = ""
        if cel65 .value is None: cel65 .value = ""
        if cel66 .value is None: cel66 .value = ""
        if cel67 .value is None: cel67 .value = ""
        if cel68 .value is None: cel68 .value = ""
        if cel69 .value is None: cel69 .value = ""
        if cel70 .value is None: cel70 .value = ""
        if cel71 .value is None: cel71 .value = ""
        if cel72 .value is None: cel72 .value = ""
        if cel73 .value is None: cel73 .value = ""
        if cel74 .value is None: cel74 .value = ""
        if cel75 .value is None: cel75 .value = ""
        if cel76 .value is None: cel76 .value = ""
        if cel77 .value is None: cel77 .value = ""
        if cel78 .value is None: cel78 .value = ""
        if cel79 .value is None: cel79 .value = ""
        if cel80 .value is None: cel80 .value = ""
        if cel81 .value is None: cel81 .value = ""
        if cel82 .value is None: cel82 .value = ""
        if cel83 .value is None: cel83 .value = ""
        if cel84 .value is None: cel84 .value = ""
        if cel85 .value is None: cel85 .value = ""
        if cel86 .value is None: cel86 .value = ""
        if cel87 .value is None: cel87 .value = ""
        if cel88 .value is None: cel88 .value = ""
        if cel89 .value is None: cel89 .value = ""
        if cel90 .value is None: cel90 .value = ""
        if cel91 .value is None: cel91 .value = ""
        if cel92 .value is None: cel92 .value = ""
        if cel93 .value is None: cel93 .value = ""
        if cel94 .value is None: cel94 .value = ""
        if cel95 .value is None: cel95 .value = ""
        if cel96 .value is None: cel96 .value = ""
        if cel97 .value is None: cel97 .value = ""
        if cel98 .value is None: cel98 .value = ""
        if cel99 .value is None: cel99 .value = ""
        if cel100 .value is None: cel100 .value = ""
        if cel101 .value is None: cel101 .value = ""
        if cel102 .value is None: cel102 .value = ""
        if cel103 .value is None: cel103 .value = ""
        if cel104 .value is None: cel104 .value = ""
        if cel105 .value is None: cel105 .value = ""
        if cel106 .value is None: cel106 .value = ""
        if cel107 .value is None: cel107 .value = ""
        if cel108 .value is None: cel108 .value = ""
        if cel109 .value is None: cel109 .value = ""
        if cel110 .value is None: cel110 .value = ""
        if cel111 .value is None: cel111 .value = ""
        if cel112 .value is None: cel112 .value = ""
        if cel113 .value is None: cel113 .value = ""
        if cel114 .value is None: cel114 .value = ""
        if cel115 .value is None: cel115 .value = ""
        if cel116 .value is None: cel116 .value = ""
        if cel117 .value is None: cel117 .value = ""
        if cel118 .value is None: cel118 .value = ""
        if cel119 .value is None: cel119 .value = ""
        if cel120 .value is None: cel120 .value = ""
        if cel121 .value is None: cel121 .value = ""
        if cel122 .value is None: cel122 .value = ""
        if cel123 .value is None: cel123 .value = ""
        if cel124 .value is None: cel124 .value = ""
        if cel125 .value is None: cel125 .value = ""
        if cel126 .value is None: cel126 .value = ""
        if cel127 .value is None: cel127 .value = ""
        if cel128 .value is None: cel128 .value = ""
        if cel129 .value is None: cel129 .value = ""
        if cel130 .value is None: cel130 .value = ""
        if cel131 .value is None: cel131 .value = ""
        if cel132 .value is None: cel132 .value = ""
        if cel133 .value is None: cel133 .value = ""
        if cel134 .value is None: cel134 .value = ""
        if cel135 .value is None: cel135 .value = ""
        if cel136 .value is None: cel136 .value = ""
        if cel137 .value is None: cel137 .value = ""
        if cel138 .value is None: cel138 .value = ""
        if cel139 .value is None: cel139 .value = ""
        if cel140 .value is None: cel140 .value = ""
        if cel141 .value is None: cel141 .value = ""
        if cel142 .value is None: cel142 .value = ""
        if cel143 .value is None: cel143 .value = ""
        if cel144 .value is None: cel144 .value = ""
        if cel145 .value is None: cel145 .value = ""
        if cel146 .value is None: cel146 .value = ""
        if cel147 .value is None: cel147 .value = ""
        if cel148 .value is None: cel148 .value = ""
        if cel149 .value is None: cel149 .value = ""
        if cel150 .value is None: cel150 .value = ""
        if cel151 .value is None: cel151 .value = ""
        if cel152 .value is None: cel152 .value = ""
        if cel153 .value is None: cel153 .value = ""
        if cel154 .value is None: cel154 .value = ""
        if cel155 .value is None: cel155 .value = ""
        if cel156 .value is None: cel156 .value = ""
        if cel157 .value is None: cel157 .value = ""
        if cel158 .value is None: cel158 .value = ""
        if cel159 .value is None: cel159 .value = ""
        if cel160 .value is None: cel160 .value = ""
        if cel161 .value is None: cel161 .value = ""
        if cel162 .value is None: cel162 .value = ""
        if cel163 .value is None: cel163 .value = ""
        if cel164 .value is None: cel164 .value = ""
        if cel165 .value is None: cel165 .value = ""
        if cel166 .value is None: cel166 .value = ""
        if cel167 .value is None: cel167 .value = ""
        if cel168 .value is None: cel168 .value = ""
        if cel169 .value is None: cel169 .value = ""
        if cel170 .value is None: cel170 .value = ""
        if cel171 .value is None: cel171 .value = ""
        if cel172 .value is None: cel172 .value = ""
        if cel173 .value is None: cel173 .value = ""
        if cel174 .value is None: cel174 .value = ""
        if cel175 .value is None: cel175 .value = ""
        if cel176 .value is None: cel176 .value = ""
        if cel177 .value is None: cel177 .value = ""
        if cel178 .value is None: cel178 .value = ""
        if cel179 .value is None: cel179 .value = ""
        if cel180 .value is None: cel180 .value = ""
        if cel181 .value is None: cel181 .value = ""
        if cel182 .value is None: cel182 .value = ""
        if cel183 .value is None: cel183 .value = ""
        if cel184 .value is None: cel184 .value = ""
        if cel185 .value is None: cel185 .value = ""
        if cel186 .value is None: cel186 .value = ""
        if cel187 .value is None: cel187 .value = ""
        if cel188 .value is None: cel188 .value = ""
        if cel189 .value is None: cel189 .value = ""
        if cel190 .value is None: cel190 .value = ""
        if cel191 .value is None: cel191 .value = ""
        if cel192 .value is None: cel192 .value = ""
        if cel193 .value is None: cel193 .value = ""
        if cel194 .value is None: cel194 .value = ""
        if cel195 .value is None: cel195 .value = ""
        if cel196 .value is None: cel196 .value = ""
        if cel197 .value is None: cel197 .value = ""
        if cel198 .value is None: cel198 .value = ""
        if cel199 .value is None: cel199 .value = ""
        if cel200 .value is None: cel200 .value = ""
        if cel201 .value is None: cel201 .value = ""
        if cel202 .value is None: cel202 .value = ""
        if cel203 .value is None: cel203 .value = ""
        if cel204 .value is None: cel204 .value = ""
        if cel205 .value is None: cel205 .value = ""
        if cel206 .value is None: cel206 .value = ""
        if cel207 .value is None: cel207 .value = ""
        if cel208 .value is None: cel208 .value = ""
        if cel209 .value is None: cel209 .value = ""
        if cel210 .value is None: cel210 .value = ""
        if cel211 .value is None: cel211 .value = ""
        if cel212 .value is None: cel212 .value = ""
        if cel213 .value is None: cel213 .value = ""
        if cel214 .value is None: cel214 .value = ""
        if cel215 .value is None: cel215 .value = ""
        if cel216 .value is None: cel216 .value = ""
        if cel217 .value is None: cel217 .value = ""
        if cel218 .value is None: cel218 .value = ""
        if cel219 .value is None: cel219 .value = ""
        if cel220 .value is None: cel220 .value = ""
        if cel221 .value is None: cel221 .value = ""
        if cel222 .value is None: cel222 .value = ""
        if cel223 .value is None: cel223 .value = ""
        if cel224 .value is None: cel224 .value = ""
        if cel225 .value is None: cel225 .value = ""
        if cel226 .value is None: cel226 .value = ""
        if cel227 .value is None: cel227 .value = ""
        if cel228 .value is None: cel228 .value = ""
        if cel229 .value is None: cel229 .value = ""
        if cel230 .value is None: cel230 .value = ""
        if cel231 .value is None: cel231 .value = ""
        if cel232 .value is None: cel232 .value = ""
        if cel233 .value is None: cel233 .value = ""
        if cel234 .value is None: cel234 .value = ""
        if cel235 .value is None: cel235 .value = ""
        if cel236 .value is None: cel236 .value = ""
        if cel237 .value is None: cel237 .value = ""
        if cel238 .value is None: cel238 .value = ""
        if cel239 .value is None: cel239 .value = ""
        if cel240 .value is None: cel240 .value = ""
        if cel241 .value is None: cel241 .value = ""
        if cel242 .value is None: cel242 .value = ""
        if cel243 .value is None: cel243 .value = ""
        if cel244 .value is None: cel244 .value = ""
        if cel245 .value is None: cel245 .value = ""
        if cel246 .value is None: cel246 .value = ""
        if cel247 .value is None: cel247 .value = ""
        if cel248 .value is None: cel248 .value = ""
        if cel249 .value is None: cel249 .value = ""
        if cel250 .value is None: cel250 .value = ""
        if cel251 .value is None: cel251 .value = ""
        if cel252 .value is None: cel252 .value = ""
        if cel253 .value is None: cel253 .value = ""
        if cel254 .value is None: cel254 .value = ""
        if cel255 .value is None: cel255 .value = ""
        if cel256 .value is None: cel256 .value = ""
        if cel257 .value is None: cel257 .value = ""
        if cel258 .value is None: cel258 .value = ""
        if cel259 .value is None: cel259 .value = ""
        if cel260 .value is None: cel260 .value = ""
        if cel261 .value is None: cel261 .value = ""
        if cel262 .value is None: cel262 .value = ""
        if cel263 .value is None: cel263 .value = ""
        if cel264 .value is None: cel264 .value = ""
        if cel265 .value is None: cel265 .value = ""
        if cel266 .value is None: cel266 .value = ""
        if cel267 .value is None: cel267 .value = ""
        if cel268 .value is None: cel268 .value = ""
        if cel269 .value is None: cel269 .value = ""
        if cel270 .value is None: cel270 .value = ""
        if cel271 .value is None: cel271 .value = ""
        if cel272 .value is None: cel272 .value = ""
        if cel273 .value is None: cel273 .value = ""
        if cel274 .value is None: cel274 .value = ""
        if cel275 .value is None: cel275 .value = ""
        if cel276 .value is None: cel276 .value = ""
        if cel277 .value is None: cel277 .value = ""
        if cel278 .value is None: cel278 .value = ""
        if cel279 .value is None: cel279 .value = ""
        if cel280 .value is None: cel280 .value = ""
        if cel281 .value is None: cel281 .value = ""
        if cel282 .value is None: cel282 .value = ""
        if cel283 .value is None: cel283 .value = ""
        if cel284 .value is None: cel284 .value = ""
        if cel285 .value is None: cel285 .value = ""
        if cel286 .value is None: cel286 .value = ""
        if cel287 .value is None: cel287 .value = ""
        if cel288 .value is None: cel288 .value = ""
        if cel289 .value is None: cel289 .value = ""
        if cel290 .value is None: cel290 .value = ""
        if cel291 .value is None: cel291 .value = ""
        if cel292 .value is None: cel292 .value = ""
        if cel293 .value is None: cel293 .value = ""
        if cel294 .value is None: cel294 .value = ""
        if cel295 .value is None: cel295 .value = ""
        if cel296 .value is None: cel296 .value = ""
        if cel297 .value is None: cel297 .value = ""
        if cel298 .value is None: cel298 .value = ""
        if cel299 .value is None: cel299 .value = ""
        if cel300 .value is None: cel300 .value = ""
        if cel301 .value is None: cel301 .value = ""
        if cel302 .value is None: cel302 .value = ""
        if cel303 .value is None: cel303 .value = ""
        if cel304 .value is None: cel304 .value = ""
        if cel305 .value is None: cel305 .value = ""
        if cel306 .value is None: cel306 .value = ""
        if cel307 .value is None: cel307 .value = ""
        if cel308 .value is None: cel308 .value = ""
        if cel309 .value is None: cel309 .value = ""
        if cel310 .value is None: cel310 .value = ""
        if cel311 .value is None: cel311 .value = ""
        if cel312 .value is None: cel312 .value = ""
        if cel313 .value is None: cel313 .value = ""
        if cel314 .value is None: cel314 .value = ""
        if cel315 .value is None: cel315 .value = ""
        if cel316 .value is None: cel316 .value = ""
        if cel317 .value is None: cel317 .value = ""
        if cel318 .value is None: cel318 .value = ""
        if cel319 .value is None: cel319 .value = ""
        if cel320 .value is None: cel320 .value = ""
        if cel321 .value is None: cel321 .value = ""
        if cel322 .value is None: cel322 .value = ""
        if cel323 .value is None: cel323 .value = ""
        if cel324 .value is None: cel324 .value = ""
        if cel325 .value is None: cel325 .value = ""
        if cel326 .value is None: cel326 .value = ""
        if cel327 .value is None: cel327 .value = ""
        if cel328 .value is None: cel328 .value = ""
        if cel329 .value is None: cel329 .value = ""
        if cel330 .value is None: cel330 .value = ""
        if cel331 .value is None: cel331 .value = ""
        if cel332 .value is None: cel332 .value = ""
        if cel333 .value is None: cel333 .value = ""
        if cel334 .value is None: cel334 .value = ""
        if cel335 .value is None: cel335 .value = ""
        if cel336 .value is None: cel336 .value = ""
        if cel337 .value is None: cel337 .value = ""
        if cel338 .value is None: cel338 .value = ""
        if cel339 .value is None: cel339 .value = ""
        if cel340 .value is None: cel340 .value = ""
        if cel341 .value is None: cel341 .value = ""
        if cel342 .value is None: cel342 .value = ""
        if cel343 .value is None: cel343 .value = ""
        if cel344 .value is None: cel344 .value = ""
        if cel345 .value is None: cel345 .value = ""
        if cel346 .value is None: cel346 .value = ""
        if cel347 .value is None: cel347 .value = ""
        if cel348 .value is None: cel348 .value = ""
        if cel349 .value is None: cel349 .value = ""
        if cel350 .value is None: cel350 .value = ""
        if cel351 .value is None: cel351 .value = ""
        if cel352 .value is None: cel352 .value = ""
        if cel353 .value is None: cel353 .value = ""
        if cel354 .value is None: cel354 .value = ""
        if cel355 .value is None: cel355 .value = ""
        if cel356 .value is None: cel356 .value = ""
        if cel357 .value is None: cel357 .value = ""
        if cel358 .value is None: cel358 .value = ""
        if cel359 .value is None: cel359 .value = ""
        if cel360 .value is None: cel360 .value = ""
        if cel361 .value is None: cel361 .value = ""
        if cel362 .value is None: cel362 .value = ""
        if cel363 .value is None: cel363 .value = ""
        if cel364 .value is None: cel364 .value = ""
        if cel365 .value is None: cel365 .value = ""
        if cel366 .value is None: cel366 .value = ""
        if cel367 .value is None: cel367 .value = ""
        if cel368 .value is None: cel368 .value = ""
        if cel369 .value is None: cel369 .value = ""
        if cel370 .value is None: cel370 .value = ""
        if cel371 .value is None: cel371 .value = ""
        if cel372 .value is None: cel372 .value = ""
        if cel373 .value is None: cel373 .value = ""
        if cel374 .value is None: cel374 .value = ""
        if cel375 .value is None: cel375 .value = ""
        if cel376 .value is None: cel376 .value = ""
        if cel377 .value is None: cel377 .value = ""
        if cel378 .value is None: cel378 .value = ""
        if cel379 .value is None: cel379 .value = ""
        if cel380 .value is None: cel380 .value = ""
        if cel381 .value is None: cel381 .value = ""
        if cel382 .value is None: cel382 .value = ""
        if cel383 .value is None: cel383 .value = ""
        if cel384 .value is None: cel384 .value = ""
        if cel385 .value is None: cel385 .value = ""
        if cel386 .value is None: cel386 .value = ""
        if cel387 .value is None: cel387 .value = ""
        if cel388 .value is None: cel388 .value = ""
        if cel389 .value is None: cel389 .value = ""
        if cel390 .value is None: cel390 .value = ""
        if cel391 .value is None: cel391 .value = ""
        if cel392 .value is None: cel392 .value = ""
        if cel393 .value is None: cel393 .value = ""
        if cel394 .value is None: cel394 .value = ""
        if cel395 .value is None: cel395 .value = ""
        if cel396 .value is None: cel396 .value = ""
        if cel397 .value is None: cel397 .value = ""
        if cel398 .value is None: cel398 .value = ""
        if cel399 .value is None: cel399 .value = ""
        if cel400 .value is None: cel400 .value = ""
        if cel401 .value is None: cel401 .value = ""
        if cel402 .value is None: cel402 .value = ""
        if cel403 .value is None: cel403 .value = ""
        if cel404 .value is None: cel404 .value = ""
        if cel405 .value is None: cel405 .value = ""
        if cel406 .value is None: cel406 .value = ""
        if cel407 .value is None: cel407 .value = ""
        if cel408 .value is None: cel408 .value = ""
        if cel409 .value is None: cel409 .value = ""
        if cel410 .value is None: cel410 .value = ""
        if cel411 .value is None: cel411 .value = ""
        if cel412 .value is None: cel412 .value = ""
        if cel413 .value is None: cel413 .value = ""
        if cel414 .value is None: cel414 .value = ""
        if cel415 .value is None: cel415 .value = ""
        if cel416 .value is None: cel416 .value = ""
        if cel417 .value is None: cel417 .value = ""
        if cel418 .value is None: cel418 .value = ""
        if cel419 .value is None: cel419 .value = ""
        if cel420 .value is None: cel420 .value = ""
        if cel421 .value is None: cel421 .value = ""
        if cel422 .value is None: cel422 .value = ""
        if cel423 .value is None: cel423 .value = ""
        if cel424 .value is None: cel424 .value = ""
        if cel425 .value is None: cel425 .value = ""
        if cel426 .value is None: cel426 .value = ""
        if cel427 .value is None: cel427 .value = ""
        if cel428 .value is None: cel428 .value = ""
        if cel429 .value is None: cel429 .value = ""
        if cel430 .value is None: cel430 .value = ""
        if cel431 .value is None: cel431 .value = ""
        if cel432 .value is None: cel432 .value = ""
        if cel433 .value is None: cel433 .value = ""
        if cel434 .value is None: cel434 .value = ""
        if cel435 .value is None: cel435 .value = ""
        if cel436 .value is None: cel436 .value = ""
        if cel437 .value is None: cel437 .value = ""
        if cel438 .value is None: cel438 .value = ""
        if cel439 .value is None: cel439 .value = ""
        if cel440 .value is None: cel440 .value = ""
        if cel441 .value is None: cel441 .value = ""
        if cel442 .value is None: cel442 .value = ""
        if cel443 .value is None: cel443 .value = ""
        if cel444 .value is None: cel444 .value = ""
        if cel445 .value is None: cel445 .value = ""
        if cel446 .value is None: cel446 .value = ""
        if cel447 .value is None: cel447 .value = ""
        if cel448 .value is None: cel448 .value = ""
        if cel449 .value is None: cel449 .value = ""
        if cel450 .value is None: cel450 .value = ""
        if cel451 .value is None: cel451 .value = ""
        if cel452 .value is None: cel452 .value = ""
        if cel453 .value is None: cel453 .value = ""
        if cel454 .value is None: cel454 .value = ""
        if cel455 .value is None: cel455 .value = ""
        if cel456 .value is None: cel456 .value = ""
        if cel457 .value is None: cel457 .value = ""
        if cel458 .value is None: cel458 .value = ""
        if cel459 .value is None: cel459 .value = ""
        if cel460 .value is None: cel460 .value = ""
        if cel461 .value is None: cel461 .value = ""
        if cel462 .value is None: cel462 .value = ""
        if cel463 .value is None: cel463 .value = ""
        if cel464 .value is None: cel464 .value = ""
        if cel465 .value is None: cel465 .value = ""
        if cel466 .value is None: cel466 .value = ""
        if cel467 .value is None: cel467 .value = ""
        if cel468 .value is None: cel468 .value = ""

        don_ent = Contas.objects.aggregate(total=Sum('donativos_Entrada'))
        don_sai = Contas.objects.aggregate(total=Sum('donativos_Saída'))
        con_ent = Contas.objects.aggregate(total=Sum('conta_Entrada'))
        con_sai = Contas.objects.aggregate(total=Sum('conta_Saída'))
        out_ent = Contas.objects.aggregate(total=Sum('outra_Entrada'))
        out_sai = Contas.objects.aggregate(total=Sum('outra_Saída'))

        don_ent_value1 = don_ent['total']
        if don_ent_value1:
            don_ent_value = float(don_ent_value1)
        else:
            don_ent_value = 0
        don_sai_value1 = don_sai['total']
        if don_sai_value1:
            don_sai_value = float(don_sai_value1)
        else:
            don_sai_value = 0   
        con_ent_value1 = con_ent['total']
        if con_ent_value1:
            con_ent_value = float(con_ent_value1)
        else:
            con_ent_value = 0    
        con_sai_value1 = con_sai['total']
        if con_sai_value1:
            con_sai_value = float(con_sai_value1)
        else:
            con_sai_value = 0    
        out_ent_value1 = out_ent['total']
        if out_ent_value1:
            out_ent_value = float(out_ent_value1)
        else:
            out_ent_value = 0    
        out_sai_value1 = out_sai['total']
        if out_sai_value1:
            out_sai_value = float(out_sai_value1)
        else:
            out_sai_value = 0   

        saldo_final_extrato_mensal2 = Gerais.objects.values('saldo_Final_do_Extrato_Mensal').last()
        if saldo_final_extrato_mensal2 is None:
            saldo_final_extrato_mensal = 0   
        else:
            saldo_final_extrato_mensal1 = saldo_final_extrato_mensal2['saldo_Final_do_Extrato_Mensal']
        if saldo_final_extrato_mensal1:
            saldo_final_extrato_mensal = float(saldo_final_extrato_mensal1)
        else:
            saldo_final_extrato_mensal = 0   

      


        remessa_betel_resolução2 = Gerais.objects.values('remessa_Enviada_para_Betel_Resolução').last()
        if remessa_betel_resolução2 is None:
            remessa_betel_resolução = 0   
        else:
            remessa_betel_resolução1 = remessa_betel_resolução2['remessa_Enviada_para_Betel_Resolução']
        if remessa_betel_resolução1:
            remessa_betel_resolução = float(remessa_betel_resolução1)
        else:
            remessa_betel_resolução = 0   

        saldo_final_extrato_betel2 = Gerais.objects.values('saldo_Final_do_Extrato_de_Betel').last()
        if saldo_final_extrato_betel2 is None:
            saldo_final_extrato_betel = 0   
        else:
            saldo_final_extrato_betel1 = saldo_final_extrato_betel2['saldo_Final_do_Extrato_de_Betel']
        if saldo_final_extrato_betel1:
            saldo_final_extrato_betel = float(saldo_final_extrato_betel1)
        else:
            saldo_final_extrato_betel = 0   

        saldo_donativos_anterior2 = Gerais.objects.values('saldo_Final_dos_Donativos_Mês_Anterior').last()
        if saldo_donativos_anterior2 is None:
            saldo_donativos_anterior = 0   
        else:
            saldo_donativos_anterior1 = saldo_donativos_anterior2['saldo_Final_dos_Donativos_Mês_Anterior']
        if saldo_donativos_anterior1:
            saldo_donativos_anterior = float(saldo_donativos_anterior1)
        else:
            saldo_donativos_anterior = 0   

    
        saldo_conta_anterior2 = Gerais.objects.values('saldo_Final_da_Conta_Bancária_Mês_Anterior').last()
        if saldo_conta_anterior2 is None:
            saldo_conta_anterior = 0   
        else:
            saldo_conta_anterior1 = saldo_conta_anterior2['saldo_Final_da_Conta_Bancária_Mês_Anterior']
        if saldo_conta_anterior1:
            saldo_conta_anterior = float(saldo_conta_anterior1)
        else:
            saldo_conta_anterior = 0   

        saldo_betel_anterior2 = Gerais.objects.values('saldo_Final_da_Conta_em_Betel_Mês_Anterior').last()
        if saldo_betel_anterior2 is None:
            saldo_betel_anterior = 0   
        else:
            saldo_betel_anterior1 = saldo_betel_anterior2['saldo_Final_da_Conta_em_Betel_Mês_Anterior']
        if saldo_betel_anterior1:
            saldo_betel_anterior = float(saldo_betel_anterior1)
        else:
            saldo_betel_anterior = 0  

        total_CF10 = Contas.objects.filter(símbolo='CF').aggregate(Sum('donativos_Entrada'))
        total_CF20 = Contas.objects.filter(símbolo='CF').aggregate(Sum('conta_Entrada'))
        total_CF30 = Contas.objects.filter(símbolo='CF').aggregate(Sum('outra_Entrada'))
        total_CFa = total_CF10[('donativos_Entrada__sum')]
        if total_CFa:
            total_CF1 = float(total_CFa)
        else:
            total_CF1 = 0  
        total_CFb = total_CF20[('conta_Entrada__sum')]
        if total_CFb:
            total_CF2 = float(total_CFb)
        else:
            total_CF2 = 0  
        total_CFc = total_CF30[('outra_Entrada__sum')]
        if total_CFc:
            total_CF3 = float(total_CFc)
        else:
            total_CF3 = 0  

        total_CF = total_CF1 + total_CF2 + total_CF3
        
        total_O10 = Contas.objects.filter(símbolo='O').aggregate(Sum('donativos_Entrada'))
        total_O20 = Contas.objects.filter(símbolo='O').aggregate(Sum('conta_Entrada'))
        total_O30 = Contas.objects.filter(símbolo='O').aggregate(Sum('outra_Entrada'))
        total_Oa = total_O10[('donativos_Entrada__sum')]
        if total_Oa:
            total_O1 = float(total_Oa)
        else:
            total_O1 = 0  
        total_Ob = total_O20[('conta_Entrada__sum')]
        if total_Ob:
            total_O2 = float(total_Ob)
        else:
            total_O2 = 0  
        total_Oc = total_O30[('outra_Entrada__sum')]
        if total_Oc:
            total_O3 = float(total_Oc)
        else:
            total_O3 = 0  

        total_O = total_O1 + total_O2 + total_O3
            

        betel = total_O + remessa_betel_resolução + total_CF
        sal_final_don = saldo_donativos_anterior + don_ent_value - don_sai_value
        sal_final_con = saldo_conta_anterior + con_ent_value - con_sai_value
        sal_final_betel = saldo_betel_anterior + out_ent_value - out_sai_value
        saldo_conta_extrato_sem_betel = saldo_final_extrato_mensal - betel
        total_total = sal_final_don + sal_final_con + sal_final_betel
        
        saldo_total_mês_anterior = saldo_donativos_anterior + saldo_conta_anterior + saldo_betel_anterior


        data_dict = {
                "congregação": cong,
                "cidade": cidade,
                "estado": estado,
                "mês": mês,
                "ano": ano,
                "data": data,
                "don_ent.total": don_ent_value,
                "don_sai.total": don_sai_value,
                "con_ent.total": con_ent_value,
                "con_sai.total": con_sai_value,
                "out_ent.total": out_ent_value,
                "out_sai.total": out_sai_value,
                "saldo_final_extrato_mensal": saldo_final_extrato_mensal,
                "betel": betel,
                "saldo_conta_extrato_sem_betel": saldo_conta_extrato_sem_betel,
                "saldo_donativos_anterior": saldo_donativos_anterior,
                "sal_final_don": sal_final_don,
                "saldo_conta_anterior": saldo_conta_anterior,
                "sal_final_con": sal_final_con,
                "saldo_betel_anterior": saldo_betel_anterior,
                "total_total": total_total,
                "saldo_final_extrato_betel": saldo_final_extrato_betel,
                "sal_final_betel": sal_final_betel,

                "1": cel1.value,
                "2": cel2.value,
                "3": cel3.value,
                "4": cel4.value,
                "5": cel5.value,
                "6": cel6.value,
                "7": cel7.value,
                "8": cel8.value,
                "9": cel9.value,
                "10": cel10.value,
                "11": cel11.value,
                "12": cel12.value,
                "13": cel13.value,
                "14": cel14.value,
                "15": cel15.value,
                "16": cel16.value,
                "17": cel17.value,
                "18": cel18.value,
                "19": cel19.value,
                "20": cel20.value,
                "21": cel21.value,
                "22": cel22.value,
                "23": cel23.value,
                "24": cel24.value,
                "25": cel25.value,
                "26": cel26.value,
                "27": cel27.value,
                "28": cel28.value,
                "29": cel29.value,
                "30": cel30.value,
                "31": cel31.value,
                "32": cel32.value,
                "33": cel33.value,
                "34": cel34.value,
                "35": cel35.value,
                "36": cel36.value,
                "37": cel37.value,
                "38": cel38.value,
                "39": cel39.value,
                "40": cel40.value,
                "41": cel41.value,
                "42": cel42.value,
                "43": cel43.value,
                "44": cel44.value,
                "45": cel45.value,
                "46": cel46.value,
                "47": cel47.value,
                "48": cel48.value,
                "49": cel49.value,
                "50": cel50.value,
                "51": cel51.value,
                "52": cel52.value,
                "53": cel53.value,
                "54": cel54.value,
                "55": cel55.value,
                "56": cel56.value,
                "57": cel57.value,
                "58": cel58.value,
                "59": cel59.value,
                "60": cel60.value,
                "61": cel61.value,
                "62": cel62.value,
                "63": cel63.value,
                "64": cel64.value,
                "65": cel65.value,
                "66": cel66.value,
                "67": cel67.value,
                "68": cel68.value,
                "69": cel69.value,
                "70": cel70.value,
                "71": cel71.value,
                "72": cel72.value,
                "73": cel73.value,
                "74": cel74.value,
                "75": cel75.value,
                "76": cel76.value,
                "77": cel77.value,
                "78": cel78.value,
                "79": cel79.value,
                "80": cel80.value,
                "81": cel81.value,
                "82": cel82.value,
                "83": cel83.value,
                "84": cel84.value,
                "85": cel85.value,
                "86": cel86.value,
                "87": cel87.value,
                "88": cel88.value,
                "89": cel89.value,
                "90": cel90.value,
                "91": cel91.value,
                "92": cel92.value,
                "93": cel93.value,
                "94": cel94.value,
                "95": cel95.value,
                "96": cel96.value,
                "97": cel97.value,
                "98": cel98.value,
                "99": cel99.value,
                "100": cel100.value,
                "101": cel101.value,
                "102": cel102.value,
                "103": cel103.value,
                "104": cel104.value,
                "105": cel105.value,
                "106": cel106.value,
                "107": cel107.value,
                "108": cel108.value,
                "109": cel109.value,
                "110": cel110.value,
                "111": cel111.value,
                "112": cel112.value,
                "113": cel113.value,
                "114": cel114.value,
                "115": cel115.value,
                "116": cel116.value,
                "117": cel117.value,
                "118": cel118.value,
                "119": cel119.value,
                "120": cel120.value,
                "121": cel121.value,
                "122": cel122.value,
                "123": cel123.value,
                "124": cel124.value,
                "125": cel125.value,
                "126": cel126.value,
                "127": cel127.value,
                "128": cel128.value,
                "129": cel129.value,
                "130": cel130.value,
                "131": cel131.value,
                "132": cel132.value,
                "133": cel133.value,
                "134": cel134.value,
                "135": cel135.value,
                "136": cel136.value,
                "137": cel137.value,
                "138": cel138.value,
                "139": cel139.value,
                "140": cel140.value,
                "141": cel141.value,
                "142": cel142.value,
                "143": cel143.value,
                "144": cel144.value,
                "145": cel145.value,
                "146": cel146.value,
                "147": cel147.value,
                "148": cel148.value,
                "149": cel149.value,
                "150": cel150.value,
                "151": cel151.value,
                "152": cel152.value,
                "153": cel153.value,
                "154": cel154.value,
                "155": cel155.value,
                "156": cel156.value,
                "157": cel157.value,
                "158": cel158.value,
                "159": cel159.value,
                "160": cel160.value,
                "161": cel161.value,
                "162": cel162.value,
                "163": cel163.value,
                "164": cel164.value,
                "165": cel165.value,
                "166": cel166.value,
                "167": cel167.value,
                "168": cel168.value,
                "169": cel169.value,
                "170": cel170.value,
                "171": cel171.value,
                "172": cel172.value,
                "173": cel173.value,
                "174": cel174.value,
                "175": cel175.value,
                "176": cel176.value,
                "177": cel177.value,
                "178": cel178.value,
                "179": cel179.value,
                "180": cel180.value,
                "181": cel181.value,
                "182": cel182.value,
                "183": cel183.value,
                "184": cel184.value,
                "185": cel185.value,
                "186": cel186.value,
                "187": cel187.value,
                "188": cel188.value,
                "189": cel189.value,
                "190": cel190.value,
                "191": cel191.value,
                "192": cel192.value,
                "193": cel193.value,
                "194": cel194.value,
                "195": cel195.value,
                "196": cel196.value,
                "197": cel197.value,
                "198": cel198.value,
                "199": cel199.value,
                "200": cel200.value,
                "201": cel201.value,
                "202": cel202.value,
                "203": cel203.value,
                "204": cel204.value,
                "205": cel205.value,
                "206": cel206.value,
                "207": cel207.value,
                "208": cel208.value,
                "209": cel209.value,
                "210": cel210.value,
                "211": cel211.value,
                "212": cel212.value,
                "213": cel213.value,
                "214": cel214.value,
                "215": cel215.value,
                "216": cel216.value,
                "217": cel217.value,
                "218": cel218.value,
                "219": cel219.value,
                "220": cel220.value,
                "221": cel221.value,
                "222": cel222.value,
                "223": cel223.value,
                "224": cel224.value,
                "225": cel225.value,
                "226": cel226.value,
                "227": cel227.value,
                "228": cel228.value,
                "229": cel229.value,
                "230": cel230.value,
                "231": cel231.value,
                "232": cel232.value,
                "233": cel233.value,
                "234": cel234.value,
                "235": cel235.value,
                "236": cel236.value,
                "237": cel237.value,
                "238": cel238.value,
                "239": cel239.value,
                "240": cel240.value,
                "241": cel241.value,
                "242": cel242.value,
                "243": cel243.value,
                "244": cel244.value,
                "245": cel245.value,
                "246": cel246.value,
                "247": cel247.value,
                "248": cel248.value,
                "249": cel249.value,
                "250": cel250.value,
                "251": cel251.value,
                "252": cel252.value,
                "253": cel253.value,
                "254": cel254.value,
                "255": cel255.value,
                "256": cel256.value,
                "257": cel257.value,
                "258": cel258.value,
                "259": cel259.value,
                "260": cel260.value,
                "261": cel261.value,
                "262": cel262.value,
                "263": cel263.value,
                "264": cel264.value,
                "265": cel265.value,
                "266": cel266.value,
                "267": cel267.value,
                "268": cel268.value,
                "269": cel269.value,
                "270": cel270.value,
                "271": cel271.value,
                "272": cel272.value,
                "273": cel273.value,
                "274": cel274.value,
                "275": cel275.value,
                "276": cel276.value,
                "277": cel277.value,
                "278": cel278.value,
                "279": cel279.value,
                "280": cel280.value,
                "281": cel281.value,
                "282": cel282.value,
                "283": cel283.value,
                "284": cel284.value,
                "285": cel285.value,
                "286": cel286.value,
                "287": cel287.value,
                "288": cel288.value,
                "289": cel289.value,
                "290": cel290.value,
                "291": cel291.value,
                "292": cel292.value,
                "293": cel293.value,
                "294": cel294.value,
                "295": cel295.value,
                "296": cel296.value,
                "297": cel297.value,
                "298": cel298.value,
                "299": cel299.value,
                "300": cel300.value,
                "301": cel301.value,
                "302": cel302.value,
                "303": cel303.value,
                "304": cel304.value,
                "305": cel305.value,
                "306": cel306.value,
                "307": cel307.value,
                "308": cel308.value,
                "309": cel309.value,
                "310": cel310.value,
                "311": cel311.value,
                "312": cel312.value,
                "313": cel313.value,
                "314": cel314.value,
                "315": cel315.value,
                "316": cel316.value,
                "317": cel317.value,
                "318": cel318.value,
                "319": cel319.value,
                "320": cel320.value,
                "321": cel321.value,
                "322": cel322.value,
                "323": cel323.value,
                "324": cel324.value,
                "325": cel325.value,
                "326": cel326.value,
                "327": cel327.value,
                "328": cel328.value,
                "329": cel329.value,
                "330": cel330.value,
                "331": cel331.value,
                "332": cel332.value,
                "333": cel333.value,
                "334": cel334.value,
                "335": cel335.value,
                "336": cel336.value,
                "337": cel337.value,
                "338": cel338.value,
                "339": cel339.value,
                "340": cel340.value,
                "341": cel341.value,
                "342": cel342.value,
                "343": cel343.value,
                "344": cel344.value,
                "345": cel345.value,
                "346": cel346.value,
                "347": cel347.value,
                "348": cel348.value,
                "349": cel349.value,
                "350": cel350.value,
                "351": cel351.value,
                "352": cel352.value,
                "353": cel353.value,
                "354": cel354.value,
                "355": cel355.value,
                "356": cel356.value,
                "357": cel357.value,
                "358": cel358.value,
                "359": cel359.value,
                "360": cel360.value,
                "361": cel361.value,
                "362": cel362.value,
                "363": cel363.value,
                "364": cel364.value,
                "365": cel365.value,
                "366": cel366.value,
                "367": cel367.value,
                "368": cel368.value,
                "369": cel369.value,
                "370": cel370.value,
                "371": cel371.value,
                "372": cel372.value,
                "373": cel373.value,
                "374": cel374.value,
                "375": cel375.value,
                "376": cel376.value,
                "377": cel377.value,
                "378": cel378.value,
                "379": cel379.value,
                "380": cel380.value,
                "381": cel381.value,
                "382": cel382.value,
                "383": cel383.value,
                "384": cel384.value,
                "385": cel385.value,
                "386": cel386.value,
                "387": cel387.value,
                "388": cel388.value,
                "389": cel389.value,
                "390": cel390.value,
                "391": cel391.value,
                "392": cel392.value,
                "393": cel393.value,
                "394": cel394.value,
                "395": cel395.value,
                "396": cel396.value,
                "397": cel397.value,
                "398": cel398.value,
                "399": cel399.value,
                "400": cel400.value,
                "401": cel401.value,
                "402": cel402.value,
                "403": cel403.value,
                "404": cel404.value,
                "405": cel405.value,
                "406": cel406.value,
                "407": cel407.value,
                "408": cel408.value,
                "409": cel409.value,
                "410": cel410.value,
                "411": cel411.value,
                "412": cel412.value,
                "413": cel413.value,
                "414": cel414.value,
                "415": cel415.value,
                "416": cel416.value,
                "417": cel417.value,
                "418": cel418.value,
                "419": cel419.value,
                "420": cel420.value,
                "421": cel421.value,
                "422": cel422.value,
                "423": cel423.value,
                "424": cel424.value,
                "425": cel425.value,
                "426": cel426.value,
                "427": cel427.value,
                "428": cel428.value,
                "429": cel429.value,
                "430": cel430.value,
                "431": cel431.value,
                "432": cel432.value,
                "433": cel433.value,
                "434": cel434.value,
                "435": cel435.value,
                "436": cel436.value,
                "437": cel437.value,
                "438": cel438.value,
                "439": cel439.value,
                "440": cel440.value,
                "441": cel441.value,
                "442": cel442.value,
                "443": cel443.value,
                "444": cel444.value,
                "445": cel445.value,
                "446": cel446.value,
                "447": cel447.value,
                "448": cel448.value,
                "449": cel449.value,
                "450": cel450.value,
                "451": cel451.value,
                "452": cel452.value,
                "453": cel453.value,
                "454": cel454.value,
                "455": cel455.value,
                "456": cel456.value,
                "457": cel457.value,
                "458": cel458.value,
                "459": cel459.value,
                "460": cel460.value,
                "461": cel461.value,
                "462": cel462.value,
                "463": cel463.value,
                "464": cel464.value,
                "465": cel465.value,
                "466": cel466.value,
                "467": cel467.value,
                "468": cel468.value,
                }

        fillpdfs.write_fillable_pdf('static/FC.pdf', 'static/FC_pronto.pdf', data_dict)

    return render(request, 'cp/imprimir_FC.html') 


def resultado(request):
    
    
    context = {
        'saldo_final_extrato_betel':saldo_final_extrato_betel,
        'sal_final_betel':sal_final_betel,
        'total_total':total_total,
        'saldo_conta_extrato_sem_betel': saldo_conta_extrato_sem_betel, 
        'sal_final_con':sal_final_con,
        'don_ent_value':don_ent_value,
        'don_sai_value':don_sai_value,
        'con_ent_value':con_ent_value,
        'con_sai_value':con_sai_value,
        'out_ent_value':out_ent_value,
        'out_sai_value':out_sai_value,
        'sal_final_don':sal_final_don,
        'total_C_TE':total_C_TE,
        'total_O':total_O,
        'total_G':total_G,
        'saldo_mês_atual':saldo_mês_atual,
        'total_betel':total_betel
        }
    
    return render(request, 'cp/resultado.html' , context )    




